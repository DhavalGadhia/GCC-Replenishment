"""
GCC Replenishment Engine - Multi-Node (production)
==================================================
Reads the five source extracts (DEPTH, RMS, SALE, SOH, WMS) straight out of
the reporting system, allocates stock from multiple source warehouses using a
country-level warehouse priority grid, and exports an Oracle-ready transfer
plan.

Run with:   streamlit run gcc_replenishment_app.py

requirements.txt:
    streamlit
    pandas
    numpy
    openpyxl
"""

import gc
import hashlib
import hmac
import secrets as _pysecrets
import os
import re
import shutil
import tempfile
import json
import time
import traceback
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font

# ---------------------------------------------------------------------------
# BUSINESS CONSTANTS
# ---------------------------------------------------------------------------
# Pack size is NOT a fixed 12. Every option's real pack size comes from the
# WMS "Pack Ratio" column. FALLBACK_PACK_SIZE is used only if a pack row
# somehow arrives without a ratio, which should not happen in practice.
FALLBACK_PACK_SIZE = 12
# Inventory reserve. Held back from the working pool, never released within a
# run - it stays in the warehouse and reappears in the next day's WMS
# snapshot. Only JAFZA carries a reserve by default; every other warehouse
# allocates from its full available quantity.
RESERVE_WH_CODE = "242211"      # JAFZA WH
DEFAULT_RESERVE_PCT = 10        # % held back at RESERVE_WH_CODE when enabled
DEFAULT_RESERVE_ON = True

# Replenishment mode. "Pack Available Qty" in the WMS extract is always in
# EACHES for both pack and loose rows (for a pack row, "Available Qty" holds
# the number of packs and Pack Ratio the units per pack). Pack quantities must
# ship as whole packs of that item's own Pack Ratio, which is not always 12.
MODE_LOOSE, MODE_PACK = "Loose", "Pack"
DEFAULT_MODE = MODE_PACK

# Warehouses in the eligibility grid, keyed by WMS "Loc Key".
GRID_WH = {
    "44324":  "DIP WH",
    "22748":  "KSA WH",
    "170001": "QAT WH",
    "242211": "JAFZA WH",
}

# WAREHOUSE-REGION ELIGIBILITY & PRIORITY (1 = tried first, absent = ineligible)
WH_REGION_PRIORITY = {
    "44324":  {"Oman": 1, "United Arab Emirates": 1},
    "22748":  {"Saudi Arabia": 1},
    "170001": {"Qatar": 1},
    "242211": {"Kuwait": 1, "Qatar": 2, "Bahrain": 1,
               "Oman": 2, "Saudi Arabia": 2, "United Arab Emirates": 2},
}

# Pack rounding threshold (% of a pack at which a remainder rounds up)
DEFAULT_PACK_THRESHOLD_PCT = {
    "Kuwait": 30, "Qatar": 30, "Bahrain": 30,
    "Oman": 50, "Saudi Arabia": 50, "United Arab Emirates": 50,
}
FALLBACK_THRESHOLD_PCT = 50

# Fallback Target Stock used when an Option + Store combination has no row in
# the DEPTH file. Without this the combination cannot be sized and is skipped
# entirely, so this is what brings those stores into the run.
DEFAULT_TARGET_STOCK = {
    "Kuwait": 24, "Qatar": 24, "Bahrain": 24,
    "Oman": 24, "Saudi Arabia": 24, "United Arab Emirates": 24,
}
FALLBACK_TARGET_STOCK = 24

# ---------------------------------------------------------------------------
# DYNAMIC TARGET STOCK (rate of sale)
#
#   ROS        = lifetime Net Sales Qty / days since First Received (per store)
#   Cover Days = SOH / ROS          -> how many days of stock the store holds
#   If Cover Days >= Target Cover  -> the store is over-covered, cut the target
#   If Cover Days <  Target Cover  -> under-covered, raise the target
#
# The increase and decrease are deliberately SEPARATE. The Overstock Failsafe
# (Raw Need = min(Base Need, Target - Pipeline)) already suppresses demand at
# well-stocked stores, so a large decrease is mostly redundant - measured at
# 61.6% of over-covered rows already sitting at zero need.
# ---------------------------------------------------------------------------
DEFAULT_COVER_DAYS = {
    "Bahrain": 56, "Kuwait": 56, "Oman": 56,
    "Qatar": 56, "Saudi Arabia": 56, "United Arab Emirates": 56,
}
DEFAULT_LEAD_DAYS = {
    "Bahrain": 8, "Kuwait": 8, "Oman": 3,
    "Qatar": 5, "Saudi Arabia": 7, "United Arab Emirates": 1,
}
# Minimum rate of sale for a line to be replenished at all. Without a SALE
# file the scope comes from LIFETIME sales, which includes lines that stopped
# selling months ago; this is the guard that excludes them. 0.02/day is about
# one unit every 50 days.
DEFAULT_MIN_ROS = 0.0

# Symmetric adjustment: the same percentage raises an under-covered target
# and lowers an over-covered one.
DEFAULT_ADJUST_PCT = 30
DEFAULT_INCREASE_PCT = DEFAULT_ADJUST_PCT
DEFAULT_DECREASE_PCT = DEFAULT_ADJUST_PCT
ROS_ADJUST_ON = True

TARGET_BASIS_DEPTH    = "Actual as per DEPTH file"
TARGET_BASIS_DEFAULT  = "Default (no ROS and no DEPTH target)"
TARGET_BASIS_UP       = "Increased by {pct:.0f}% of target stock"
TARGET_BASIS_DOWN     = "Reduced by {pct:.0f}% of target stock"

# Only these columns are pulled from each workbook - reading all 44/56 columns
# of a 55 MB extract is the single biggest avoidable cost.
FILE_SPECS = {
    "DEPTH": {
        "signature": {"OPTION", "New Depth"},
        "header_row": 3,                       # two blank rows precede the header
        "cols": ["OPTION", "Store code", "New Depth"],
    },
    "RMS": {
        "signature": {"Open Order Qty", "From Loc Key"},
        "header_row": 1,
        "cols": ["From Loc Key", "To Loc Key", "To Loc Type", "Option", "Open Order Qty",
                 "Pack Indicator"],
    },
    "SALE": {
        "signature": {"Net Sales Qty", "Location Code"},
        "header_row": 1,
        "cols": ["Country", "Location Code", "Location", "Item Style Code",
                 "Item Color", "Net Sales Qty"],
    },
    "WMS": {
        "signature": {"Pack Available Qty", "Loc Key"},
        "header_row": 1,
        "cols": ["Country", "Loc Key", "Location", "Option", "Pack Available Qty",
                 "Uda Product Type", "Item Key", "Pack Indicator", "Pack Ratio"],
    },
    # SOH is a wide pivot and is handled by its own reader.
    "SOH": {"signature": {"Pack PhysicalQty", "option"}, "header_row": 4, "cols": None},
}


# Bump whenever a READER changes shape (not just FILE_SPECS), otherwise the
# cached frames from the previous reader keep being served.
READER_VERSION = "soh-ros-v3"


def _spec_fingerprint():
    """
    Hash of the file-reading specs. This is passed into the cached loader so
    that changing which columns are pulled AUTOMATICALLY invalidates the
    cache. Without it, st.cache_data keys only on the uploaded bytes and the
    loader's own source - so a change to FILE_SPECS (a module-level global)
    would silently keep serving frames parsed under the OLD column list.
    """
    payload = json.dumps(
        {k: {"cols": v["cols"], "header_row": v["header_row"],
             "signature": sorted(v["signature"])}
         for k, v in FILE_SPECS.items()},
        sort_keys=True) + "|" + READER_VERSION
    return hashlib.md5(payload.encode()).hexdigest()[:12]


SPEC_VERSION = _spec_fingerprint()


def build_country_priority():
    out = {}
    for wh, regions in WH_REGION_PRIORITY.items():
        for country, rank in regions.items():
            out.setdefault(country, []).append((rank, wh))
    return {c: [wh for _, wh in sorted(v)] for c, v in out.items()}


COUNTRY_PRIORITY = build_country_priority()

# Colour palette cycled for warehouses added at run time
WH_PALETTE = ["#2F6FED", "#0EA5A0", "#E0A008", "#8B5CF6", "#E05252",
              "#0891B2", "#DB2777", "#65A30D"]


def routing_rows_from(priority_map, countries):
    """Matrix rows: one per country, one column per warehouse code."""
    rows = []
    for c in countries:
        row = {"Country": c}
        chain = priority_map.get(c, [])
        for code in sorted({w for ch in priority_map.values() for w in ch}):
            row[code] = (chain.index(code) + 1) if code in chain else None
        rows.append(row)
    return rows


def priority_from_matrix(df, wh_codes):
    """
    Turn the edited matrix back into {country: [codes ranked]}. A blank cell
    means the warehouse is not eligible for that country. Ties are broken by
    the warehouse's column order so the result is always deterministic.
    """
    out = {}
    for _, r in df.iterrows():
        country = str(r["Country"]).strip()
        ranked = []
        for pos, code in enumerate(wh_codes):
            v = pd.to_numeric(r.get(code), errors="coerce")
            if pd.notna(v) and v > 0:
                ranked.append((float(v), pos, code))
        if ranked:
            out[country] = [c for _, _, c in sorted(ranked)]
    return out

_S = lambda s: s.astype(str).str.strip()


# ---------------------------------------------------------------------------
# ROUNDING
# ---------------------------------------------------------------------------
def round_to_pack(qty, pack_size=FALLBACK_PACK_SIZE, threshold_pct=FALLBACK_THRESHOLD_PCT) -> int:
    """Scalar form. Remainder at/above pack_size x threshold rounds up."""
    if pd.isna(qty) or qty <= 0:
        return 0
    r = qty % pack_size
    if r == 0:
        return int(qty)
    return int(qty - r) if r < pack_size * threshold_pct / 100.0 else int(qty - r + pack_size)


def round_to_pack_vec(qty, threshold_pct, pack_size):
    """
    Vectorised rounding. pack_size is PER ROW - each option is rounded to its
    own pack size taken from the WMS Pack Ratio, not to a fixed 12. A pack
    size of 1 means eaches, i.e. no rounding at all (loose replenishment).
    """
    q = np.asarray(qty, dtype=float)
    t = np.asarray(threshold_pct, dtype=float)
    p = np.maximum(np.asarray(pack_size, dtype=float), 1.0)
    r = np.mod(q, p)
    cut = p * t / 100.0
    rounded = np.where(r == 0, q, np.where(r < cut, q - r, q - r + p))
    return np.where(q > 0, rounded, 0).astype(np.int64)


# ---------------------------------------------------------------------------
# FILE INTAKE
# ---------------------------------------------------------------------------
def _is_csvlike(name):
    return str(name).lower().endswith((".csv", ".tsv", ".txt"))


def _seek0(src):
    """No-op for a path; rewinds a file-like object."""
    if hasattr(src, "seek"):
        src.seek(0)
    return src


def _norm(v):
    """
    Normalise a header cell for matching: case, spacing, underscores and
    punctuation are all ignored. 'Net Sales Qty', 'net_sales_qty' and
    'NET SALES QTY ' all collapse to 'netsalesqty'.
    """
    if v is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(v).strip().lower())


# Accepted spellings per canonical field. An EXACT normalised match on the
# canonical name always wins; these aliases are only tried when the canonical
# name is absent. Deliberately conservative - a wrong column silently
# producing a plausible plan is far worse than a clear error. In particular
# 'Available Qty' is NOT an alias for 'Pack Available Qty' (one is packs, the
# other eaches) and 'Net Sales Amt' is never treated as a quantity.
FIELD_ALIASES = {
    "OPTION": {"option", "optioncode", "optionid"},
    "Option": {"option", "optioncode", "optionid"},
    "option": {"option", "optioncode", "optionid"},
    "Store code": {"storecode", "storecodes", "storeno", "storenumber",
                   "storeid", "loccode", "locationcd"},
    # NOTE: bare "dept"/"depth" are deliberately NOT aliases - SOH carries a
    # "Dept" (department) column and would be misread as a DEPTH file.
    "New Depth": {"newdepth", "newdept", "targetstock", "targetdepth",
                  "maxcap", "maxcapacity"},
    "From Loc Key": {"fromlockey", "fromlocationkey", "fromloc", "fromlocation",
                     "sourcelockey", "sourcelocation"},
    "To Loc Key": {"tolockey", "tolocationkey", "toloc", "tolocation",
                   "destlockey", "destinationlocation"},
    "To Loc Type": {"toloctype", "tolocationtype", "destlocationtype"},
    "Open Order Qty": {"openorderqty", "openorderquantity", "openqty",
                       "openorder"},
    "Pack Indicator": {"packindicator", "packind", "ispack", "packflag"},
    "Pack Ratio": {"packratio", "packsize", "packqty", "unitsperpack"},
    "Country": {"country", "countryname", "countrydesc"},
    "Location Code": {"locationcode", "loccode", "storecode", "locationcd",
                      "locationkey", "lockey"},
    "Location": {"location", "locationname", "storename", "locname"},
    "Item Style Code": {"itemstylecode", "stylecode", "style"},
    "Item Color": {"itemcolor", "itemcolour", "color", "colour", "clr"},
    "Net Sales Qty": {"netsalesqty", "netsalesquantity", "netsaleqty",
                      "netsoldqty", "totalsoldqty", "soldqty"},
    "Loc Key": {"lockey", "locationkey", "warehousecode", "whcode",
                "warehousekey"},
    "Pack Available Qty": {"packavailableqty", "packavailqty",
                           "packavailablequantity"},
    "Uda Product Type": {"udaproducttype", "producttype"},
    "Item Key": {"itemkey", "itembarcode", "barcode"},
    "Pack PhysicalQty": {"packphysicalqty", "packphysicalquantity",
                         "physicalqty"},
}

SCAN_ROWS = 25        # how deep to hunt for a header row
SCAN_SHEETS = 8       # how many worksheets to consider


def _alias_set(field):
    return FIELD_ALIASES.get(field, set()) | {_norm(field)}


def _resolve_columns(header_cells, wanted):
    """
    Map canonical field names to column positions in a header row.
    Returns (colmap, missing). Exact normalised matches are taken first so an
    alias can never displace a column that is genuinely present.
    """
    norm = [_norm(c) for c in header_cells]
    taken, colmap = set(), {}

    for field in wanted:                       # pass 1 - exact
        target = _norm(field)
        for i, n in enumerate(norm):
            if n and n == target and i not in taken:
                colmap[field] = i
                taken.add(i)
                break

    for field in wanted:                       # pass 2 - accepted aliases
        if field in colmap:
            continue
        opts = _alias_set(field)
        for i, n in enumerate(norm):
            if n and n in opts and i not in taken:
                colmap[field] = i
                taken.add(i)
                break

    return colmap, [f for f in wanted if f not in colmap]


def _score_header(cells, kind):
    """How many of this file kind's required columns a candidate row resolves."""
    colmap, _ = _resolve_columns(cells, FILE_SPECS[kind]["cols"] or [])
    return len(colmap)


def _signature_ok(cells, kind):
    """
    Signature match for file identification - EXACT normalised names only.
    Aliases are used to locate columns once the file kind is known, but never
    to decide the kind: a loose alias could otherwise make one extract
    masquerade as another.
    """
    norm = {_norm(c) for c in cells if c is not None}
    return {_norm(f) for f in FILE_SPECS[kind]["signature"]} <= norm


def _xlsx_rows(path, sheet=None, max_row=SCAN_ROWS):
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(max_row=max_row, values_only=True)]
    finally:
        wb.close()


def locate_xlsx_table(path, kind):
    """
    Find where this file kind's table actually starts: which worksheet, which
    header row, and which column holds each field. Nothing about position is
    assumed - the header may sit on any row of any sheet, and the columns may
    be in any order.
    Returns (sheet_name, header_row_index, colmap).
    """
    spec = FILE_SPECS[kind]
    wanted = spec["cols"] or []
    wb = load_workbook(path, read_only=True)
    sheets = list(wb.sheetnames)[:SCAN_SHEETS]
    wb.close()

    best = None
    for sheet in sheets:
        rows = _xlsx_rows(path, sheet)
        for r, cells in enumerate(rows):
            if not any(c is not None for c in cells):
                continue
            colmap, missing = _resolve_columns(cells, wanted)
            if missing:
                continue
            score = len(colmap)
            if best is None or score > best[0]:
                best = (score, sheet, r, colmap)
            break                       # first complete header on this sheet wins
    if best is None:
        # Nothing complete - report the closest attempt so the message is useful
        detail = ""
        for sheet in sheets:
            for cells in _xlsx_rows(path, sheet):
                colmap, missing = _resolve_columns(cells, wanted)
                if colmap and missing:
                    detail = (f" Closest match on sheet '{sheet}' was missing: "
                              f"{', '.join(missing)}.")
                    break
            if detail:
                break
        raise ValueError(
            f"{kind}: could not find a header row containing the required "
            f"columns ({', '.join(wanted)}) in the first {SCAN_ROWS} rows of "
            f"the first {len(sheets)} sheet(s).{detail}")
    return best[1], best[2], best[3]


def identify_file(buf, name=""):
    """
    Work out which of the five extracts this is, by looking for its signature
    columns anywhere in the opening rows of any sheet - not at a fixed
    position.
    """
    if _is_csvlike(name):
        try:
            head = _csv_head(buf, nrows=SCAN_ROWS)
        except Exception:
            return None
        hits = []
        for kind in FILE_SPECS:
            for i in range(len(head)):
                if _signature_ok(list(head.iloc[i]), kind):
                    hits.append((_score_header(list(head.iloc[i]), kind), kind))
                    break
        return max(hits)[1] if hits else None
    try:
        wb = load_workbook(buf, read_only=True)
        sheets = list(wb.sheetnames)[:SCAN_SHEETS]
        wb.close()
        hits = []
        for kind in FILE_SPECS:
            for sheet in sheets:
                found = False
                for cells in _xlsx_rows(buf, sheet):
                    if _signature_ok(cells, kind):
                        hits.append((_score_header(cells, kind), kind))
                        found = True
                        break
                if found:
                    break
        return max(hits)[1] if hits else None
    except Exception:
        return None


def read_flat(buf, kind):
    """Read the needed columns from a flat sheet, wherever they happen to sit."""
    sheet, hdr_row, colmap = locate_xlsx_table(buf, kind)
    cols = FILE_SPECS[kind]["cols"]
    wb = load_workbook(buf, read_only=True)
    try:
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        for _ in range(hdr_row + 1):
            next(it)
        ii = [colmap[c] for c in cols]
        width = max(ii) + 1
        rows = [tuple(r[i] if i < len(r) else None for i in ii)
                for r in it if len(r) >= 1]
    finally:
        wb.close()
    return pd.DataFrame(rows, columns=cols)


def _csv_head(src, nrows=SCAN_ROWS):
    _seek0(src)
    return pd.read_csv(src, header=None, nrows=nrows, dtype=str,
                       sep=None, engine="python", on_bad_lines="skip")


def _find_csv_header(src, kind):
    """Locate the header row and column positions in a delimited file."""
    head = _csv_head(src)
    wanted = FILE_SPECS[kind]["cols"] or []
    for i in range(len(head)):
        cells = [None if pd.isna(c) else c for c in head.iloc[i]]
        colmap, missing = _resolve_columns(cells, wanted)
        if not missing:
            return i, colmap
    raise ValueError(
        f"{kind}: could not find a header row containing the required columns "
        f"({', '.join(wanted)}) in the first {SCAN_ROWS} rows.")


def read_flat_csv(src, kind):
    """Same resilient column location, for a delimited text file."""
    hdr, colmap = _find_csv_header(src, kind)
    cols = FILE_SPECS[kind]["cols"]
    _seek0(src)
    df = pd.read_csv(src, skiprows=hdr, sep=None, engine="python",
                     dtype=str, on_bad_lines="skip", header=0)
    out = pd.DataFrame({c: df.iloc[:, colmap[c]] for c in cols})
    return out.reset_index(drop=True)


def _locate_soh_blocks(rows):
    """
    Work out the shape of the SOH crosstab without assuming it has exactly
    four header rows. Finds the metric header row (the one carrying both
    'option' and 'Pack PhysicalQty'), treats the rows above it as the
    per-store metadata band, and derives the item-attribute columns from the
    leading columns where that band is empty.
    Returns (hdr_idx, opt_col, blocks).
    """
    hdr_idx = None
    for i, cells in enumerate(rows[:SCAN_ROWS]):
        norm = {_norm(c) for c in cells if c is not None}
        if {"option", "packphysicalqty"} <= norm:
            hdr_idx = i
            break
    if hdr_idx is None:
        raise ValueError(
            "SOH: could not find the header row carrying both 'option' and "
            f"'Pack PhysicalQty' in the first {SCAN_ROWS} rows.")

    hdr = rows[hdr_idx]
    meta = rows[:hdr_idx]
    if not meta:
        raise ValueError(
            "SOH: found the metric header row but no store metadata rows above "
            "it - this does not look like the wide per-store export.")

    opt_col = next(i for i, c in enumerate(hdr) if _norm(c) == "option")

    n = len(hdr)
    code_row = meta[-1]                                  # most specific = store code

    # Identify the COUNTRY row by content, never by position. Exports vary in
    # how many banner/summary rows sit above the store metadata, and anchoring
    # to meta[0] silently picked up a numeric totals row when one was added.
    known = {_norm(c) for c in COUNTRY_PRIORITY}

    def _country_score(row):
        vals = [v for v in row if v is not None]
        if not vals:
            return 0.0
        return sum(1 for v in vals if _norm(v) in known) / len(vals)

    scored = [(_country_score(r), i) for i, r in enumerate(meta)]
    best_score, best_i = max(scored)
    ctry_row = meta[best_i] if best_score >= 0.5 else meta[0]

    # Store name = the lowest metadata row that is neither the country row nor
    # the code row.
    name_candidates = [i for i in range(len(meta))
                       if i != best_i and i != len(meta) - 1]
    name_row = meta[name_candidates[-1]] if name_candidates else meta[-1]

    def cell(row, i):
        return row[i] if i < len(row) else None

    attr_end = 0
    while attr_end < n and all(cell(m, attr_end) is None for m in meta):
        attr_end += 1

    blocks, j = [], attr_end
    while j < n:
        code, start = cell(code_row, j), j
        while j < n and cell(code_row, j) == code:
            j += 1
        metrics = {_norm(hdr[k]): k for k in range(start, min(j, len(hdr)))}
        if "packphysicalqty" in metrics:
            # Net sales and first-received drive the rate-of-sale calculation.
            # Both are optional: an older SOH export without them still works,
            # the engine simply falls back to the DEPTH target.
            blocks.append((cell(ctry_row, start), cell(name_row, start), code,
                           metrics["packphysicalqty"],
                           metrics.get("netsalesqty"),
                           metrics.get("firstrecieveddatestore",
                                       metrics.get("firstreceiveddatestore"))))
    if not blocks:
        raise ValueError(
            "SOH: found the header row but no per-store 'Pack PhysicalQty' "
            "columns beneath the store metadata.")
    return hdr_idx, opt_col, blocks


def read_soh_pivot(buf):
    """SOH wide crosstab (xlsx), with its shape detected rather than assumed."""
    head = _xlsx_rows(buf, None, max_row=SCAN_ROWS)
    hdr_idx, opt_col, blocks = _locate_soh_blocks(head)

    wb = load_workbook(buf, read_only=True)
    recs = []
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        for _ in range(hdr_idx + 1):
            next(it)
        for r in it:
            if opt_col >= len(r):
                continue
            opt = r[opt_col]
            if opt is None:
                continue
            for country, store_name, code, ci, ci_s, ci_d in blocks:
                q = r[ci] if ci < len(r) else None
                sold = r[ci_s] if ci_s is not None and ci_s < len(r) else None
                frd = r[ci_d] if ci_d is not None and ci_d < len(r) else None
                if q or sold or frd:
                    recs.append((opt, code, store_name, country, q, sold, frd))
    finally:
        wb.close()
    return pd.DataFrame(recs, columns=["Option", "Store code", "Location",
                                       "Country", "SOH", "NetSalesQty",
                                       "FirstReceived"])


def read_soh_pivot_csv(buf):
    """SOH wide crosstab from a delimited text file."""
    _seek0(buf)
    raw = pd.read_csv(buf, header=None, sep=None, engine="python",
                      dtype=str, on_bad_lines="skip")
    if len(raw) < 3:
        raise ValueError("SOH file has too few rows to be the wide pivot export.")
    grid = [[None if pd.isna(v) else v for v in raw.iloc[i]]
            for i in range(min(len(raw), SCAN_ROWS))]
    hdr_idx, opt_col, blocks = _locate_soh_blocks(grid)

    recs = []
    for row in raw.iloc[hdr_idx + 1:].itertuples(index=False):
        opt = row[opt_col] if opt_col < len(row) else None
        if opt is None or (isinstance(opt, float) and pd.isna(opt)):
            continue
        for country, store_name, code, ci, ci_s, ci_d in blocks:
            q = row[ci] if ci < len(row) else None
            sold = row[ci_s] if ci_s is not None and ci_s < len(row) else None
            frd = row[ci_d] if ci_d is not None and ci_d < len(row) else None
            ok = lambda v: v is not None and str(v) not in ("nan", "")
            if ok(q) or ok(sold) or ok(frd):
                recs.append((opt, code, store_name, country, q, sold, frd))
    return pd.DataFrame(recs, columns=["Option", "Store code", "Location",
                                       "Country", "SOH", "NetSalesQty",
                                       "FirstReceived"])


def upload_fingerprint(uploads):
    """
    Short cache key built from Streamlit's own per-upload file_id (falling back
    to name+size). Used INSTEAD of hashing the file bytes: st.cache_data hashes
    every argument it can see, and hashing ~140 MB of workbooks on each rerun
    both costs seconds and keeps a second copy of the bytes alive.
    """
    h = hashlib.md5()
    for f in uploads:
        fid = getattr(f, "file_id", None) or f"{f.name}:{getattr(f, 'size', 0)}"
        h.update(str(fid).encode())
    return h.hexdigest()[:16]


@st.cache_data(show_spinner=False, max_entries=1)
def load_sources(_uploads, cache_key, spec_version):
    """
    _uploads: Streamlit UploadedFile objects. The leading underscore tells
    st.cache_data NOT to hash this argument - caching is keyed on cache_key
    (the upload fingerprint) and spec_version instead. Each file's bytes are
    materialised one at a time and released immediately, so peak memory is
    the largest single file rather than the sum of all five.
    Returns (frames dict, per-file read seconds, warnings).
    """
    frames, timings, warnings = {}, {}, []
    for f in _uploads:
        name = f.name
        # Stream the upload to a temp file and parse from disk. Streamlit
        # already holds the upload in memory; calling getvalue() would hold a
        # SECOND full copy, and on a 1 GB container five large extracts is
        # what tips the process over. openpyxl and pandas both read lazily
        # from a path, so peak memory becomes the parse itself, not the bytes.
        try:
            f.seek(0)
        except Exception:
            pass
        suffix = os.path.splitext(name)[1] or ".xlsx"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        try:
            shutil.copyfileobj(f, tmp, length=1024 * 1024)
        except TypeError:                     # object without a file interface
            tmp.write(f.getvalue())
        tmp.close()
        path = tmp.name
        try:
            kind = identify_file(path, name)
            if kind is None:
                warnings.append(
                    f"'{name}' was not recognised as DEPTH, RMS, SALE, SOH or WMS.")
                continue
            t0 = time.time()
            try:
                if _is_csvlike(name):
                    df = read_soh_pivot_csv(path) if kind == "SOH" else read_flat_csv(path, kind)
                else:
                    df = read_soh_pivot(path) if kind == "SOH" else read_flat(path, kind)
            except Exception as e:
                warnings.append(
                    f"'{name}' ({kind}) could not be read: {type(e).__name__}: {e}")
                continue
            frames[kind] = (pd.concat([frames[kind], df], ignore_index=True)
                            if kind in frames else df)
            timings[kind] = timings.get(kind, 0) + (time.time() - t0)
        finally:
            # Always remove the temp file, including on an unrecognised or
            # unreadable upload - otherwise a failed run leaves 100+ MB behind.
            try:
                os.unlink(path)
            except OSError:
                pass
            gc.collect()

    return frames, timings, warnings


# ---------------------------------------------------------------------------
# ENGINE
# ---------------------------------------------------------------------------
def prepare(frames, active_wh, reserve_map=None, wh_names=None):
    """Normalise the five extracts into the tables the allocator consumes."""
    missing = [k for k in ("DEPTH", "SOH", "WMS") if k not in frames]
    if missing:
        raise ValueError(
            "Missing required file(s): " + ", ".join(missing)
            + ". Upload the DEPTH, SOH and WMS extracts. RMS is optional but "
              "strongly recommended - without it, in-transit stock is unknown "
              "and the engine will over-order. SALE is optional: when supplied "
              "it decides which lines are in scope, otherwise the SOH lifetime "
              "sales are used."
        )

    depth = frames["DEPTH"]
    sale = frames.get("SALE")          # optional
    soh, wms = frames["SOH"], frames["WMS"]
    rms = frames.get("RMS")
    warnings = []

    depth = depth.assign(
        Option=_S(depth["OPTION"]), Store=_S(depth["Store code"]),
        Target=pd.to_numeric(depth["New Depth"], errors="coerce").fillna(0),
    ).groupby(["Option", "Store"], as_index=False)["Target"].max()

    # Reporting exports often carry a grand-total row above the data. It has
    # no style code or location but a very large quantity, so it must be
    # dropped before anything is aggregated or it doubles reported sales.
    if sale is not None:
      _key_blank = (sale["Item Style Code"].isna() | sale["Location Code"].isna()
                  | _S(sale["Item Style Code"]).str.lower().isin(["", "nan", "none", "total"])
                  | _S(sale["Location Code"]).str.lower().isin(["", "nan", "none", "total"]))
      if _key_blank.any():
        _dropped_qty = pd.to_numeric(
            sale.loc[_key_blank, "Net Sales Qty"], errors="coerce").fillna(0).sum()
        warnings.append(
            f"{int(_key_blank.sum())} SALE row(s) had no style code or location "
            f"(typically a grand-total row) and were excluded, removing "
            f"{_dropped_qty:,.0f} units of phantom demand.")
        sale = sale[~_key_blank]
      sale = sale.assign(
        Option=_S(sale["Item Style Code"]) + _S(sale["Item Color"]),
        Store=_S(sale["Location Code"]),
        Qty=pd.to_numeric(sale["Net Sales Qty"], errors="coerce").fillna(0),
      )
    _soh_keyed = soh.assign(Store=_S(soh["Store code"]))
    country_map = pd.concat(
        ([sale[["Store", "Country"]]] if sale is not None else [])
        + [_soh_keyed[["Store", "Country"]]]
    ).dropna().drop_duplicates("Store")

    # Store code -> store name. SALE and SOH both carry it; SALE wins on
    # conflict because it is the transactional source.
    store_names = pd.concat(
        ([sale[["Store", "Location"]].rename(columns={"Location": "Store Name"})]
         if sale is not None else [])
        + [_soh_keyed[["Store", "Location"]].rename(columns={"Location": "Store Name"})]
    ).dropna().drop_duplicates("Store")

    # Option -> budget category, from the WMS extract. An Option can appear on
    # many warehouse rows, so take the first non-blank value per Option.
    if "Uda Product Type" in wms.columns:
        bc = wms.assign(Option=_S(wms["Option"]),
                        **{"Budget Category": _S(wms["Uda Product Type"])})
        bc = bc[bc["Budget Category"].notna() & (bc["Budget Category"] != "")
                & (bc["Budget Category"].str.lower() != "none")]
        budget_cat = bc.drop_duplicates("Option")[["Option", "Budget Category"]]
    else:
        warnings.append("WMS file has no 'Uda Product Type' column - Budget Category "
                        "will be blank.")
        budget_cat = pd.DataFrame(columns=["Option", "Budget Category"])

    # Option -> real pack size, from the WMS Pack Ratio on pack rows. Verified
    # consistent within an option in the source data.
    pr = wms.assign(Option=_S(wms["Option"]),
                    R=pd.to_numeric(wms.get("Pack Ratio"), errors="coerce"))
    pr = pr[(_S(pr["Pack Indicator"]).str.upper() == "Y") & pr["R"].notna() & (pr["R"] > 0)]
    pack_size_map = (pr.groupby("Option")["R"].max().astype(int).to_dict()
                     if len(pr) else {})

    # SCOPE. With a SALE file, a line is in scope only if it sold in that
    # (recent) window. Without one, the SOH lifetime sales stand in - which is
    # far more permissive, because "has ever sold" is not "is selling now".
    # min_ros is the guard that replaces the recency SALE used to provide.
    if sale is not None:
        sales = sale.groupby(["Option", "Store"], as_index=False)["Qty"].sum()
        sales = sales[sales["Qty"] > 0]
        scope_basis = "SALE file (recent period)"
    else:
        sales = _soh_keyed.assign(
            Option=_S(_soh_keyed["Option"]),
            Qty=pd.to_numeric(_soh_keyed.get("NetSalesQty"), errors="coerce").fillna(0),
        ).groupby(["Option", "Store"], as_index=False)["Qty"].sum()
        sales = sales[sales["Qty"] > 0]
        scope_basis = "SOH lifetime sales"
        warnings.append(
            "No SALE file uploaded - scope is taken from SOH lifetime sales. "
            "A line is included if it has EVER sold, not if it is selling now; "
            "use the minimum rate of sale setting to exclude dormant lines.")

    soh_norm = soh.assign(
        Option=_S(soh["Option"]), Store=_S(soh["Store code"]),
        SOH=pd.to_numeric(soh["SOH"], errors="coerce").fillna(0),
    )
    soh_ag = soh_norm.groupby(["Option", "Store"], as_index=False)["SOH"].sum()

    # Rate of sale, if the SOH export carries lifetime sales and the first
    # received date. Older exports without those columns simply yield no ROS
    # and the engine falls back to the DEPTH target.
    if {"NetSalesQty", "FirstReceived"} <= set(soh_norm.columns):
        ros = soh_norm.copy()
        ros["Sold"] = pd.to_numeric(ros["NetSalesQty"], errors="coerce")
        ros["First"] = pd.to_datetime(ros["FirstReceived"], errors="coerce")
        ros = ros.dropna(subset=["First"])
        ros["Days"] = (pd.Timestamp.now().normalize() - ros["First"]).dt.days
        ros = ros[ros["Days"] > 0]
        ros["ROS"] = ros["Sold"].clip(lower=0) / ros["Days"]
        ros = (ros.groupby(["Option", "Store"], as_index=False)
                  .agg(**{"ROS": ("ROS", "max"), "Days On Sale": ("Days", "max"),
                          "Lifetime Sold": ("Sold", "max")}))
    else:
        warnings.append(
            "SOH file has no 'Net Sales Qty' / 'First recieved date store' columns - "
            "rate-of-sale adjustment is inactive and DEPTH targets are used as-is.")
        ros = pd.DataFrame(columns=["Option", "Store", "ROS", "Days On Sale",
                                    "Lifetime Sold"])

    if rms is not None and len(rms):
        rms = rms.assign(
            Option=_S(rms["Option"]), To=_S(rms["To Loc Key"]),
            Open=pd.to_numeric(rms["Open Order Qty"], errors="coerce").fillna(0),
        )
        in_transit = (rms[rms["To Loc Type"] == "S"]
                      .groupby(["Option", "To"], as_index=False)["Open"].sum()
                      .rename(columns={"To": "Store", "Open": "InTransit"}))
        # Warehouse-bound movements, kept as their own bucket so the audit
        # trail can say whether a line was paid for by stock already sitting
        # in the warehouse or by stock still on its way in.
        wh_bound = rms[(rms["To Loc Type"] == "W") & (rms["To"].isin(active_wh))].copy()
        wh_bound["From"] = _S(wh_bound["From Loc Key"])
        wh_bound["Mode"] = np.where(
            _S(wh_bound.get("Pack Indicator", pd.Series("N", index=wh_bound.index))
               ).str.upper() == "Y", "P", "L")
        inbound = (wh_bound.groupby(["Option", "To", "Mode"], as_index=False)["Open"].sum()
                   .rename(columns={"To": "WH", "Open": "Inbound"}))
        inbound_src = (wh_bound.groupby(["Option", "To", "Mode"])["From"]
                       .agg(lambda x: ", ".join(sorted({(wh_names or GRID_WH).get(v, v)
                                                        for v in x})))
                       .reset_index().rename(columns={"To": "WH", "From": "Inbound From"}))
        inbound = inbound.merge(inbound_src, on=["Option", "WH", "Mode"], how="left")
    else:
        warnings.append(
            "No RMS file uploaded - in-transit and yet-to-dispatch stock is being "
            "treated as zero, which will overstate demand."
        )
        in_transit = pd.DataFrame(columns=["Option", "Store", "InTransit"])
        inbound = pd.DataFrame(columns=["Option", "WH", "Mode", "Inbound", "Inbound From"])

    wms_ag = wms.assign(
        Option=_S(wms["Option"]), WH=_S(wms["Loc Key"]),
        Avail=pd.to_numeric(wms["Pack Available Qty"], errors="coerce").fillna(0),
        Ratio=pd.to_numeric(wms.get("Pack Ratio"), errors="coerce"),
    )
    wms_ag["Mode"] = np.where(
        _S(wms_ag["Pack Indicator"]).str.upper() == "Y", "P", "L")
    wms_ag = wms_ag[wms_ag["WH"].isin(active_wh)]

    # One barcode per Option x WH x mode. Pack rows carry exactly one Item Key;
    # loose rows carry one per size, so the barcode holding the most stock is
    # used as the representative and the full list is kept for the audit trail.
    keyed = wms_ag[wms_ag["Avail"] > 0].sort_values("Avail", ascending=False)
    item_keys = (keyed.groupby(["Option", "WH", "Mode"], as_index=False)
                 .agg(**{"Item Key": ("Item Key", "first"),
                         "Item Keys (all)": ("Item Key",
                                             lambda x: ", ".join(_S(pd.Series(x)).unique()[:8])),
                         "Pack Ratio": ("Ratio", "first")}))

    wms_ag = wms_ag.groupby(["Option", "WH", "Mode"], as_index=False)["Avail"].sum()
    wms_ag = wms_ag.merge(item_keys, on=["Option", "WH", "Mode"], how="left")

    if len(inbound):
        # A warehouse can have inbound stock for an Option it currently holds
        # none of, so this must be an outer join or that stock is invisible.
        wms_ag = wms_ag.merge(inbound, on=["Option", "WH", "Mode"], how="outer")
    else:
        wms_ag["Inbound"] = 0.0
        wms_ag["Inbound From"] = ""
    for c, fill in [("Avail", 0), ("Inbound", 0), ("Inbound From", ""),
                    ("Item Key", ""), ("Item Keys (all)", "")]:
        wms_ag[c] = wms_ag[c].fillna(fill)
    # Pack rows must have a ratio; fall back only if the extract omits it.
    wms_ag["Pack Ratio"] = np.where(
        wms_ag["Mode"] == "P",
        pd.to_numeric(wms_ag["Pack Ratio"], errors="coerce").fillna(FALLBACK_PACK_SIZE),
        1)
    wms_ag["Opening"] = wms_ag["Avail"] + wms_ag["Inbound"]

    # Reserve is per-warehouse: a warehouse absent from reserve_map keeps 0%,
    # so its entire available quantity is allocatable. It is applied at the
    # same rate to both buckets.
    rmap = reserve_map or {}
    wms_ag["Reserve %"] = wms_ag["WH"].map(rmap).fillna(0.0)
    keep = 1 - wms_ag["Reserve %"] / 100.0
    wms_ag["Pool Own"] = (wms_ag["Avail"] * keep).round().astype(np.int64)
    wms_ag["Pool Inbound"] = (wms_ag["Inbound"] * keep).round().astype(np.int64)
    wms_ag["Working Pool"] = wms_ag["Pool Own"] + wms_ag["Pool Inbound"]
    wms_ag["Reserve Held"] = wms_ag["Opening"] - wms_ag["Working Pool"]

    return (sales, depth, soh_ag, in_transit, wms_ag, country_map,
            store_names, budget_cat, pack_size_map, ros, scope_basis, warnings)


def build_universe(sales, depth, soh_ag, in_transit, country_map,
                   pack_thresholds, default_targets=None,
                   pack_size_map=None, mode=DEFAULT_MODE,
                   cover_days=None, increase_pct=None, decrease_pct=None,
                   ros_adjust=ROS_ADJUST_ON, ros_tbl=None, min_ros=DEFAULT_MIN_ROS):
    """
    Scope = option x store with sales > 0. Target Stock comes from DEPTH where
    a row exists; otherwise the country's default Target Stock is applied, so
    a selling store is never skipped just because DEPTH has no row for it.
    """
    # Country must be attached BEFORE the target is resolved, because the
    # fallback target is country-specific.
    u = sales.merge(country_map, on="Store", how="left")
    u["Country"] = u["Country"].fillna("UNKNOWN")

    u = u.merge(depth, on=["Option", "Store"], how="left")

    dflt = default_targets or DEFAULT_TARGET_STOCK
    missing = u["Target"].isna()
    n_defaulted = int(missing.sum())
    u["Target Source"] = np.where(missing, "Default grid", "DEPTH file")
    u["Base Target"] = u["Target"]
    u.loc[missing, "Target"] = (
        u.loc[missing, "Country"].map(dflt).fillna(FALLBACK_TARGET_STOCK)
    )
    u["Base Target"] = u["Target"]

    u = (u.merge(soh_ag, on=["Option", "Store"], how="left")
           .merge(in_transit, on=["Option", "Store"], how="left"))
    if ros_tbl is not None and len(ros_tbl):
        u = u.merge(ros_tbl, on=["Option", "Store"], how="left")
    else:
        u["ROS"] = np.nan
    u[["SOH", "InTransit"]] = u[["SOH", "InTransit"]].fillna(0)

    # ---- minimum rate of sale ------------------------------------------
    # Applied only where a ROS figure exists, so lines with no ROS are never
    # silently dropped by it.
    n_slow = 0
    if min_ros and min_ros > 0 and "ROS" in u.columns:
        too_slow = u["ROS"].notna() & (u["ROS"] < float(min_ros))
        n_slow = int(too_slow.sum())
        u = u[~too_slow].copy()

    # ---- dynamic adjustment by rate of sale ----------------------------
    inc = float(DEFAULT_INCREASE_PCT if increase_pct is None else increase_pct)
    dec = float(DEFAULT_DECREASE_PCT if decrease_pct is None else decrease_pct)
    cov = cover_days or {c: DEFAULT_COVER_DAYS.get(c, 56) + DEFAULT_LEAD_DAYS.get(c, 0)
                         for c in DEFAULT_COVER_DAYS}

    u["Target Cover Days"] = u["Country"].map(cov)
    # Recomputed from a stored column, not from the earlier `missing` mask:
    # the minimum-ROS filter above may have dropped rows, so that mask no
    # longer aligns with u.
    was_default = u["Target Source"].eq("Default grid").to_numpy()
    u["Target Stock Basis"] = np.where(was_default, TARGET_BASIS_DEFAULT,
                                       TARGET_BASIS_DEPTH)

    if ros_adjust and "ROS" in u.columns:
        has_ros = u["ROS"].notna() & u["Target Cover Days"].notna()
        # Cover Days = SOH / ROS. Stock on hand with no sales at all is
        # infinite cover, which counts as over-covered.
        soh_pos = u["SOH"].clip(lower=0)
        with np.errstate(divide="ignore", invalid="ignore"):
            cover = np.where(u["ROS"] > 0, soh_pos / u["ROS"].replace(0, np.nan), np.inf)
        u["Cover Days"] = np.where(has_ros, cover, np.nan)

        over = has_ros & (u["Cover Days"] >= u["Target Cover Days"])
        under = has_ros & (u["Cover Days"] < u["Target Cover Days"])
        u.loc[over, "Target"] = (u.loc[over, "Base Target"] * (1 - dec / 100)).round()
        u.loc[under, "Target"] = (u.loc[under, "Base Target"] * (1 + inc / 100)).round()
        u["Target"] = u["Target"].clip(lower=0)
        u.loc[over, "Target Stock Basis"] = TARGET_BASIS_DOWN.format(pct=dec)
        u.loc[under, "Target Stock Basis"] = TARGET_BASIS_UP.format(pct=inc)
    else:
        u["Cover Days"] = np.nan

    # Open Order Qty already covers in-transit AND yet-to-dispatch, so there is
    # no separate yet-to-dispatch term to add here.
    u["Total Pipeline"] = u["SOH"] + u["InTransit"]
    u["Base Need"] = np.where(u["Qty"] > 0, u["Target"], 0)
    u["Overstock Failsafe"] = u["Target"] - u["Total Pipeline"]
    u["Raw Need"] = np.minimum(u["Base Need"], u["Overstock Failsafe"])
    stockout = u["Total Pipeline"] <= 0
    u.loc[stockout, "Raw Need"] = u.loc[stockout, "Target"]

    thr = pack_thresholds or DEFAULT_PACK_THRESHOLD_PCT
    u["Pack Threshold %"] = u["Country"].map(thr).fillna(FALLBACK_THRESHOLD_PCT)

    # Demand is sized in the unit it will actually ship in:
    #   Pack  -> the option's own Pack Ratio from WMS (never a fixed 12)
    #   Loose -> eaches, so pack size 1 and no rounding applied
    if mode == MODE_PACK:
        u["Pack Size Used"] = (u["Option"].map(pack_size_map or {})
                               .fillna(1).astype(int))
    else:
        u["Pack Size Used"] = 1
    u["Rounded Need"] = round_to_pack_vec(u["Raw Need"], u["Pack Threshold %"],
                                          u["Pack Size Used"])
    return u, n_defaulted, n_slow


def allocate(u, wms_ag, mode=DEFAULT_MODE, country_priority=None, wh_names=None):
    """
    Rank every store by Total Sold Qty (highest first) within its Option, then
    walk that Option's country priority chain and take the full need from the
    first eligible warehouse that can cover it.

    A run is EITHER pack or loose, never both:
      Pack  - one barcode per option, ships only in whole packs of that
              option's own Pack Ratio. The need was already sized to that
              ratio in build_universe, so it is taken as-is.
      Loose - individual barcodes, ships in eaches, no pack rounding.

    Within the chosen form each warehouse has two buckets, drawn own stock
    first and then inbound in-transit, so the audit trail can say which one
    paid for the line.
    """
    form_code = "P" if mode == MODE_PACK else "L"
    form_name = "Pack" if mode == MODE_PACK else "Loose"
    cprio = country_priority or COUNTRY_PRIORITY
    names = wh_names or GRID_WH

    keys = list(zip(wms_ag["Option"].to_numpy(), wms_ag["WH"].to_numpy(),
                    wms_ag["Mode"].to_numpy()))
    pool_own = dict(zip(keys, wms_ag["Pool Own"].to_numpy().astype(int)))
    pool_inb = dict(zip(keys, wms_ag["Pool Inbound"].to_numpy().astype(int)))
    inb_from = dict(zip(keys, wms_ag["Inbound From"].to_numpy()))
    item_key = dict(zip(keys, wms_ag["Item Key"].to_numpy()))
    ratios = dict(zip(keys, wms_ag["Pack Ratio"].to_numpy().astype(int)))

    u = u.sort_values(["Option", "Qty"], ascending=[True, False], kind="mergesort")
    u["Sales Rank"] = u.groupby("Option").cumcount() + 1

    options = u["Option"].to_numpy()
    countries = u["Country"].to_numpy()
    needs = u["Rounded Need"].to_numpy()

    n = len(u)
    alloc = np.zeros(n, dtype=np.int64)
    from_own = np.zeros(n, dtype=np.int64)
    from_inb = np.zeros(n, dtype=np.int64)
    npacks = np.zeros(n, dtype=np.int64)
    pratio = np.zeros(n, dtype=np.int64)
    source = np.empty(n, dtype=object)
    status = np.empty(n, dtype=object)
    remark = np.empty(n, dtype=object)
    bkey = np.empty(n, dtype=object)

    for i in range(n):
        need = needs[i]
        if need <= 0:
            source[i] = ""; status[i] = "No demand"; remark[i] = ""; bkey[i] = ""
            continue
        chain = cprio.get(countries[i])
        if not chain:
            source[i] = ""; status[i] = "No eligible warehouse"; remark[i] = ""
            bkey[i] = ""
            continue

        placed = False
        for rank, wh in enumerate(chain, start=1):
            key = (options[i], wh, form_code)
            own = pool_own.get(key, 0)
            inb = pool_inb.get(key, 0)
            if own + inb < need:
                continue

            take_own = min(own, need)
            take_inb = need - take_own
            if take_own:
                pool_own[key] = own - take_own
            if take_inb:
                pool_inb[key] = inb - take_inb

            ratio = max(1, int(ratios.get(key, FALLBACK_PACK_SIZE))) if form_code == "P" else 1
            alloc[i] = need
            from_own[i] = take_own
            from_inb[i] = take_inb
            pratio[i] = ratio if form_code == "P" else 0
            npacks[i] = need // ratio if form_code == "P" else 0
            source[i] = wh
            bkey[i] = item_key.get(key, "")

            whn = names.get(wh, wh)
            src = inb_from.get(key) or "another warehouse"
            tag = f"P{rank}"
            qdesc = (f"{npacks[i]} pack(s) of {ratio} ({need} units)"
                     if form_code == "P" else f"{need} loose units")
            if take_inb == 0:
                bucket = f"{whn} WMS inventory alone"
            elif take_own == 0:
                bucket = f"in-transit inbound from {src} ({whn} had no own stock)"
            else:
                bucket = (f"{take_own} from {whn} WMS inventory + {take_inb} "
                          f"from in-transit inbound from {src}")

            if rank == 1:
                status[i] = f"Filled {tag} - {form_name.lower()}"
                remark[i] = f"{tag}: {qdesc} fulfilled from {bucket}"
            else:
                p1 = names.get(chain[0], chain[0])
                status[i] = f"Filled {tag} - {form_name.lower()} fallback from {whn}"
                remark[i] = (f"{tag}: {p1} could not cover this line; {qdesc} "
                             f"fulfilled direct from {bucket}")
            placed = True
            break

        if not placed:
            source[i] = ""; bkey[i] = ""
            status[i] = "Unfilled - no eligible warehouse had stock"
            chain_names = " > ".join(names.get(w, w) for w in chain)
            remark[i] = (f"No {form_name.lower()} stock (own or inbound) at any "
                         f"eligible warehouse: {chain_names}")

    u["Allocated Qty"] = alloc
    u["Qty from Own Stock"] = from_own
    u["Qty from In-Transit"] = from_inb
    u["Refill Form"] = np.where(alloc > 0, form_name, "")
    u["Item Key"] = bkey
    u["Pack Ratio Used"] = pratio
    u["Packs Shipped"] = npacks
    u["Source WH Code"] = source
    u["Source WH"] = pd.Series(source, index=u.index).map(names).fillna("")
    u["Fulfilment Status"] = status
    u["Source Remark"] = remark
    return u, pool_own, pool_inb


def run_engine(frames, active_wh, pack_thresholds=None, default_targets=None,
               reserve_map=None, mode=DEFAULT_MODE, country_priority=None,
               wh_names=None, user_email=None, cover_days=None,
               increase_pct=None, decrease_pct=None, ros_adjust=ROS_ADJUST_ON,
               min_ros=DEFAULT_MIN_ROS, progress_cb=None):
    def report(f, label):
        if progress_cb:
            progress_cb(f, label)

    t0 = time.time()
    report(0.10, "Normalising extracts and mapping stores to countries")
    (sales, depth, soh_ag, in_transit, wms_ag, cmap, store_names, budget_cat,
     pack_size_map, ros_tbl, scope_basis, warns) = prepare(
         frames, active_wh, reserve_map, wh_names)

    report(0.35, "Calculating pipeline, failsafes and pack rounding")
    u, n_defaulted, n_slow = build_universe(sales, depth, soh_ag, in_transit,
                                            cmap, pack_thresholds, default_targets,
                                            pack_size_map, mode, cover_days,
                                            increase_pct, decrease_pct, ros_adjust,
                                            ros_tbl, min_ros)
    if n_slow:
        warns.append(
            f"{n_slow:,} option x store line(s) were excluded for selling below the "
            f"minimum rate of sale ({min_ros:g}/day).")

    report(0.60, f"Allocating across warehouses - {u['Option'].nunique():,} options")
    detail, pool_own, pool_inb = allocate(u, wms_ag, mode, country_priority, wh_names)

    # Descriptive lookups for the output and audit trail
    detail = (detail.merge(store_names, on="Store", how="left")
                    .merge(budget_cat, on="Option", how="left"))
    detail["Store Name"] = detail["Store Name"].fillna("")
    detail["Budget Category"] = detail["Budget Category"].fillna("")

    # These columns repeat a handful of values across 100k+ rows, so category
    # dtype cuts the audit trail's memory by roughly half.
    for c in ("Country", "Source WH", "Source WH Code", "Fulfilment Status",
              "Refill Form", "Target Source", "Budget Category"):
        if c in detail.columns:
            detail[c] = detail[c].astype("category")

    # Stamp who ran it - first column of the audit trail, so any line can be
    # traced back to the planner who generated it.
    # Surface the target and the rule that produced it near the front of the
    # audit trail - it is the first thing a planner checks on a surprising line.
    for col, pos in (("Target Stock Basis", 0), ("Target Stock", 0)):
        if col == "Target Stock" and "Target" in detail.columns:
            detail.insert(pos, "Target Stock", detail["Target"])
        elif col in detail.columns:
            detail.insert(pos, col, detail.pop(col))

    detail.insert(0, "User", user_email or "unknown")
    detail.insert(1, "Run At", time.strftime("%Y-%m-%d %H:%M:%S"))

    report(0.90, "Building Oracle transfer plan")
    hit = detail[detail["Allocated Qty"] > 0]
    final = pd.DataFrame({
        "FROM LOCATION": hit["Source WH Code"].to_numpy(),
        "TO LOCATION": hit["Store"].to_numpy(),
        "ITEM": hit["Option"].to_numpy(),
        "QUANTITY": hit["Allocated Qty"].to_numpy().astype(int),
        "BUDGET CATEGORY": hit["Budget Category"].to_numpy(),
        "STORE NAME": hit["Store Name"].to_numpy(),
        "ITEM KEY": hit["Item Key"].to_numpy(),
    })

    shipped = (final.groupby("FROM LOCATION")["QUANTITY"].sum()
               if len(final) else pd.Series(dtype=float))
    wh_summary = wms_ag.groupby("WH", as_index=False).agg(
        Opening=("Opening", "sum"),
        **{"Own Stock": ("Avail", "sum"),
           "Inbound In-Transit": ("Inbound", "sum"),
           "Working Pool": ("Working Pool", "sum"),
           "Reserve Held": ("Reserve Held", "sum"),
           "Reserve %": ("Reserve %", "first")})
    wh_summary["WH Name"] = wh_summary["WH"].map(wh_names or GRID_WH)
    wh_summary["Shipped"] = wh_summary["WH"].map(shipped).fillna(0).astype(int)
    wh_summary["Unused Pool"] = wh_summary["Working Pool"] - wh_summary["Shipped"]
    wh_summary["Pool Utilisation %"] = (
        wh_summary["Shipped"] / wh_summary["Working Pool"].replace(0, np.nan) * 100
    ).round(1)
    shipped_own = (detail.groupby("Source WH Code")["Qty from Own Stock"].sum()
                   if len(detail) else pd.Series(dtype=float))
    shipped_inb = (detail.groupby("Source WH Code")["Qty from In-Transit"].sum()
                   if len(detail) else pd.Series(dtype=float))
    wh_summary["Shipped from Own"] = wh_summary["WH"].map(shipped_own).fillna(0).astype(int)
    wh_summary["Shipped from In-Transit"] = wh_summary["WH"].map(shipped_inb).fillna(0).astype(int)
    wh_summary = wh_summary[["WH", "WH Name", "Own Stock", "Inbound In-Transit", "Opening",
                             "Reserve %", "Reserve Held", "Working Pool", "Shipped",
                             "Shipped from Own", "Shipped from In-Transit",
                             "Unused Pool", "Pool Utilisation %"]]

    # Release the big intermediates before returning. On a 1 GB container the
    # peak, not the average, is what kills a run.
    del u, sales, soh_ag, in_transit, pool_own, pool_inb
    gc.collect()

    report(1.0, "Complete")
    return final, detail, wh_summary, warns, n_defaulted, time.time() - t0


# ---------------------------------------------------------------------------
# EXPORT
# ---------------------------------------------------------------------------
def to_excel(sheets, sheet_name="Sheet1"):
    """sheets: a DataFrame, or a dict of {sheet_name: DataFrame} for a
    multi-sheet workbook (used to split Pack and Loose)."""
    if isinstance(sheets, pd.DataFrame):
        sheets = {sheet_name: sheets}
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])
            ws = writer.sheets[name[:31]]
            for cell in ws[1]:
                cell.font = Font(bold=True)
            n = min(len(df), 200)
            sample = df.head(n)
            for i, col in enumerate(df.columns, start=1):
                letter = ws.cell(row=1, column=i).column_letter
                mx = sample[col].astype(str).map(len).max() if n else 0
                ws.column_dimensions[letter].width = min(
                    max(len(str(col)), int(mx) if pd.notna(mx) else 8) + 4, 40)
    return out.getvalue()


def summarise(final, detail, wh_summary, n_defaulted):
    need = int(detail["Rounded Need"].sum())
    got = int(detail["Allocated Qty"].sum())
    return {
        "units": int(final["QUANTITY"].sum()) if len(final) else 0,
        "lines": len(final),
        "items": final["ITEM"].nunique() if len(final) else 0,
        "stores": final["TO LOCATION"].nunique() if len(final) else 0,
        "need": need, "got": got,
        "fill": (got / need * 100) if need else 0.0,
        "reserve": int(wh_summary["Reserve Held"].sum()),
        "fallback": int(detail["Fulfilment Status"].str.contains("fallback", na=False).sum()),
        "unfilled": int(detail["Fulfilment Status"].str.startswith("Unfilled", na=False).sum()),
        "unmet_units": int(detail.loc[
            detail["Fulfilment Status"].str.startswith("Unfilled", na=False), "Rounded Need"].sum()),
        "defaulted": n_defaulted,
        "defaulted_units": int(detail.loc[detail["Target Source"] == "Default grid", "Allocated Qty"].sum()),
        "wh_used": final["FROM LOCATION"].nunique() if len(final) else 0,
    }


# ---------------------------------------------------------------------------
# AUTHENTICATION
#
# Passwords are NEVER stored in readable form. Each one is salted and put
# through PBKDF2-HMAC-SHA256 (260,000 rounds); only the salt and the resulting
# hash are written to disk, and sign-in re-derives the hash to compare. A
# leaked credentials file therefore does not reveal anybody's password.
#
# STORAGE - PLEASE READ:
# The credentials file lives on the container's local disk. Streamlit
# Community Cloud gives an app NO persistent disk: the file survives normal
# use, but is wiped whenever the app reboots, sleeps and wakes, or you push a
# new commit. After that everyone is treated as a first-time user and sets a
# password again. To make credentials durable, either point AUTH_FILE at real
# persistent storage, or pre-seed users through Streamlit secrets (below),
# which survives restarts because it is held by the platform, not the disk.
# ---------------------------------------------------------------------------
AUTH_FILE = os.environ.get("GCC_AUTH_FILE", ".auth/users.json")
PBKDF2_ROUNDS = 260_000
MIN_PASSWORD_LEN = 8


def _auth_path():
    p = os.path.abspath(AUTH_FILE)
    os.makedirs(os.path.dirname(p), exist_ok=True)
    return p


def _secret_block(name, default=None):
    """Read st.secrets[...] without exploding when no secrets file exists."""
    try:
        return st.secrets[name]
    except Exception:
        return default


def load_users():
    """
    Credential registry: {email: {salt, hash, rounds, created, last_login}}.
    Seeded from st.secrets["auth"]["users"] when present, so pre-provisioned
    accounts survive a container restart that wipes the local file.
    """
    users = {}
    seed = (_secret_block("auth", {}) or {})
    try:
        seeded = dict(seed.get("users", {}) or {})
    except Exception:
        seeded = {}
    for email, rec in seeded.items():
        if isinstance(rec, dict) and rec.get("hash") and rec.get("salt"):
            users[email.strip().lower()] = dict(rec)

    try:
        with open(_auth_path(), "r", encoding="utf-8") as fh:
            disk = json.load(fh)
        if isinstance(disk, dict):
            users.update({k.strip().lower(): v for k, v in disk.items()})
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        pass
    return users


def save_users(users):
    try:
        with open(_auth_path(), "w", encoding="utf-8") as fh:
            json.dump(users, fh, indent=1)
        return True
    except OSError:
        return False


def hash_password(password, salt=None):
    salt = salt or _pysecrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt.encode("utf-8"), PBKDF2_ROUNDS)
    return salt, digest.hex()


def verify_password(password, rec):
    if not rec or not rec.get("salt") or not rec.get("hash"):
        return False
    rounds = int(rec.get("rounds", PBKDF2_ROUNDS))
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), rec["salt"].encode("utf-8"), rounds)
    return hmac.compare_digest(digest.hex(), rec["hash"])


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[A-Za-z]{2,}$")


def email_allowed(email):
    """
    Registration allowlist. Without one, anyone who reaches the app could
    self-register, which is a door rather than a lock - so when neither an
    email nor a domain list is configured the app says so in the UI.
    Configure in .streamlit/secrets.toml:
        [auth]
        allowed_emails = ["planner@yourco.com"]
        allowed_domains = ["yourco.com"]
    """
    auth = _secret_block("auth", {}) or {}
    try:
        emails = {str(e).strip().lower() for e in (auth.get("allowed_emails") or [])}
        domains = {str(d).strip().lower().lstrip("@")
                   for d in (auth.get("allowed_domains") or [])}
    except Exception:
        emails, domains = set(), set()
    if not emails and not domains:
        return True, "open"
    email = email.strip().lower()
    if email in emails:
        return True, "listed"
    if email.rsplit("@", 1)[-1] in domains:
        return True, "domain"
    return False, "blocked"


def password_problem(pw, confirm):
    if len(pw) < MIN_PASSWORD_LEN:
        return f"Password must be at least {MIN_PASSWORD_LEN} characters."
    if pw != confirm:
        return "The two passwords do not match."
    if pw.strip() == "":
        return "Password cannot be blank."
    return None


def _login_shell(step_no, step_label, body_fn):
    """Renders the login card. body_fn draws the step-specific widgets."""
    st.html(f"""
    <div class="auth-wrap">
      <div class="auth-card">
        <div class="auth-brand">
          <span class="auth-dot"></span>
          <span class="auth-name">GCC REPLENISHMENT ENGINE</span>
        </div>
        <div class="auth-steps">
          <span class="as {'on' if step_no >= 1 else ''}">1 &middot; Identify</span>
          <span class="as-rail"></span>
          <span class="as {'on' if step_no >= 2 else ''}">2 &middot; {step_label}</span>
        </div>
      </div>
    </div>
    """)
    body_fn()


def render_login():
    """Two-step sign-in. Returns only when the visitor is authenticated."""
    st.session_state.setdefault("auth_stage", "email")
    st.session_state.setdefault("auth_email", "")

    users = load_users()
    stage = st.session_state.auth_stage
    email = st.session_state.auth_email

    left, mid, right = st.columns([1, 2, 1])
    with mid:
        # ---------- step 1: email ----------
        if stage == "email":
            _login_shell(1, "Password", lambda: None)
            st.markdown('<p class="auth-lead">Enter your work email to continue.</p>',
                        unsafe_allow_html=True)
            entered = st.text_input("Email address", key="auth_email_input",
                                    placeholder="you@company.com")
            if st.button("Continue", type="primary", use_container_width=True):
                e = (entered or "").strip().lower()
                if not EMAIL_RE.match(e):
                    st.error("That does not look like a valid email address.")
                else:
                    ok, reason = email_allowed(e)
                    if not ok:
                        st.error("This email is not authorised to use the platform. "
                                 "Ask your administrator to add it.")
                    else:
                        st.session_state.auth_email = e
                        st.session_state.auth_stage = (
                            "password" if e in users else "create")
                        st.rerun()
            ok, reason = email_allowed("probe@example.com")
            if reason == "open":
                st.warning(
                    "No allowlist is configured, so any email address can register. "
                    "Set allowed_emails or allowed_domains under [auth] in "
                    "Streamlit secrets before going live.")

        # ---------- step 2a: returning user ----------
        elif stage == "password":
            _login_shell(2, "Password", lambda: None)
            st.markdown(f'<p class="auth-lead">Welcome back, <b>{email}</b></p>',
                        unsafe_allow_html=True)
            pw = st.text_input("Password", type="password", key="auth_pw_input")
            c1, c2 = st.columns([3, 1])
            with c1:
                if st.button("Sign in", type="primary", use_container_width=True):
                    if verify_password(pw or "", users.get(email)):
                        rec = users[email]
                        rec["last_login"] = time.strftime("%Y-%m-%d %H:%M:%S")
                        users[email] = rec
                        save_users(users)
                        st.session_state.auth_user = email
                        st.session_state.auth_stage = "done"
                        st.rerun()
                    else:
                        st.error("Incorrect password. Please try again.")
            with c2:
                if st.button("Back", use_container_width=True):
                    st.session_state.auth_stage = "email"
                    st.rerun()

        # ---------- step 2b: first-time user ----------
        elif stage == "create":
            _login_shell(2, "Create password", lambda: None)
            st.markdown(
                f'<p class="auth-lead">First sign-in for <b>{email}</b>. '
                f'Choose a password.</p>', unsafe_allow_html=True)
            pw1 = st.text_input("New password", type="password", key="auth_new1")
            pw2 = st.text_input("Confirm new password", type="password", key="auth_new2")
            st.caption(f"At least {MIN_PASSWORD_LEN} characters. Stored only as a "
                       f"salted PBKDF2 hash - never in readable form.")
            c1, c2 = st.columns([3, 1])
            with c1:
                if st.button("Create password and sign in", type="primary",
                             use_container_width=True):
                    problem = password_problem(pw1 or "", pw2 or "")
                    if problem:
                        st.error(problem)
                    else:
                        salt, digest = hash_password(pw1)
                        users[email] = {"salt": salt, "hash": digest,
                                        "rounds": PBKDF2_ROUNDS,
                                        "created": time.strftime("%Y-%m-%d %H:%M:%S"),
                                        "last_login": time.strftime("%Y-%m-%d %H:%M:%S")}
                        if not save_users(users):
                            st.warning("Password set for this session, but the "
                                       "credentials file could not be written - it "
                                       "will not persist.")
                        st.session_state.auth_user = email
                        st.session_state.auth_stage = "done"
                        st.rerun()
            with c2:
                if st.button("Back", use_container_width=True):
                    st.session_state.auth_stage = "email"
                    st.rerun()

    st.stop()


def require_login():
    if not st.session_state.get("auth_user"):
        render_login()
    return st.session_state.auth_user


# ---------------------------------------------------------------------------
# THEME
# ---------------------------------------------------------------------------
def inject_theme():
    st.html("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@500;600;700&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@400;500;700&display=swap');
    :root{ --ink:#0C1222; --ink-2:#141C33; --muted:#7A88A6; --line:#E4E8F0;
           --accent:#2F6FED; --teal:#0EA5A0; --amber:#E0A008; --rose:#E05252; --surface:#fff; }
    html, body, [class*="css"]{ font-family:'Inter',sans-serif; }
    .main .block-container{ padding-top:1.2rem; padding-bottom:4rem; max-width:1240px; }
    h1,h2,h3,h4{ font-family:'Space Grotesk',sans-serif !important; color:var(--ink); }
    footer{ visibility:hidden; }

    @keyframes riseIn{ from{opacity:0; transform:translateY(18px) scale(.985);} to{opacity:1; transform:none;} }
    @keyframes fadeIn{ from{opacity:0} to{opacity:1} }
    @keyframes laneFlow{ 0%{background-position:0% 50%} 100%{background-position:200% 50%} }
    @keyframes barGrow{ from{width:0%} }
    @keyframes pulseDot{ 0%,100%{opacity:.35;transform:scale(.85)} 50%{opacity:1;transform:scale(1.15)} }
    .rise{ animation:riseIn .55s cubic-bezier(.22,.68,.35,1) both; }
    .d1{animation-delay:.05s}.d2{animation-delay:.13s}.d3{animation-delay:.21s}
    .d4{animation-delay:.29s}.d5{animation-delay:.37s}.d6{animation-delay:.45s}
    @media (prefers-reduced-motion: reduce){ .rise,.hero,.kpi,.lane-track,.lane-dot{animation:none !important} }

    .hero{ position:relative; overflow:hidden;
      background: radial-gradient(1200px 300px at 12% -40%, rgba(47,111,237,.45), transparent 60%),
                  linear-gradient(118deg,#0C1222 0%,#1B2547 58%,#243163 100%);
      border-radius:18px; padding:30px 34px 26px; margin-bottom:26px;
      box-shadow:0 18px 44px rgba(12,18,34,.28); animation:fadeIn .6s ease both; }
    .hero-eyebrow{ font-family:'JetBrains Mono',monospace; font-size:.7rem; letter-spacing:.16em;
      text-transform:uppercase; color:#8DA2E0; }
    .hero-title{ font-family:'Space Grotesk',sans-serif; font-size:2.1rem; font-weight:700;
      color:#fff; margin:8px 0 4px; letter-spacing:-.02em; }
    .hero-sub{ font-size:.94rem; color:#B9C6EA; margin:0; }
    .lane{ display:flex; align-items:center; gap:12px; margin-top:20px; flex-wrap:wrap; }
    .lane-node{ font-family:'JetBrains Mono',monospace; font-size:.72rem; color:#E9EEFC;
      background:rgba(255,255,255,.10); border:1px solid rgba(255,255,255,.22);
      padding:6px 13px; border-radius:999px; white-space:nowrap; }
    .lane-track{ flex:1; min-width:60px; height:3px; border-radius:3px;
      background:linear-gradient(90deg,rgba(255,255,255,.12) 0%,var(--accent) 25%,var(--teal) 50%,rgba(255,255,255,.12) 75%);
      background-size:200% 100%; animation:laneFlow 3.4s linear infinite; }
    .lane-dot{ width:7px;height:7px;border-radius:50%;background:var(--teal);
      animation:pulseDot 1.8s ease-in-out infinite; }

    .step{ font-family:'JetBrains Mono',monospace; font-size:.7rem; letter-spacing:.12em;
      text-transform:uppercase; color:var(--accent); font-weight:600; margin:2px 0 4px; }
    .sec-title{ font-family:'Space Grotesk',sans-serif; font-weight:600; font-size:1.14rem;
      color:var(--ink); margin:0; }
    .sec-sub{ font-size:.84rem; color:var(--muted); margin-top:3px; }

    .kpi{ position:relative; background:var(--surface); border:1px solid var(--line);
      border-radius:14px; padding:18px 20px 16px; height:100%;
      box-shadow:0 1px 2px rgba(12,18,34,.05), 0 12px 26px rgba(12,18,34,.07);
      transition:transform .22s ease, box-shadow .22s ease; overflow:hidden; }
    .kpi:hover{ transform:translateY(-3px); box-shadow:0 18px 38px rgba(12,18,34,.13); }
    .kpi::before{ content:''; position:absolute; left:0; top:0; bottom:0; width:4px; background:var(--accent); }
    .kpi.teal::before{background:var(--teal)} .kpi.amber::before{background:var(--amber)}
    .kpi.rose::before{background:var(--rose)}
    .kpi-eyebrow{ font-family:'JetBrains Mono',monospace; font-size:.65rem; letter-spacing:.12em;
      text-transform:uppercase; color:var(--muted); }
    .kpi-value{ font-family:'Space Grotesk',sans-serif; font-size:2rem; font-weight:700;
      color:var(--ink); line-height:1.08; margin:7px 0 3px; font-variant-numeric:tabular-nums; }
    .kpi-label{ font-size:.85rem; color:var(--ink-2); }
    .kpi-cap{ font-size:.74rem; color:var(--muted); margin-top:5px; }
    .meter{ height:5px; border-radius:4px; background:#EDF1F7; margin-top:11px; overflow:hidden; }
    .meter > span{ display:block; height:100%; border-radius:4px;
      background:linear-gradient(90deg,var(--teal),var(--accent));
      animation:barGrow 1.1s cubic-bezier(.22,.68,.35,1) both; }
    .alert{ background:#FDF6E7; border:1px solid #F0DCA4; border-radius:11px;
      padding:12px 17px; font-size:.87rem; color:#7A5B05; margin:14px 0 4px; }
    .alert b{ color:#5B4200; }
    .stDataFrame{ border-radius:11px; overflow:hidden; }

    /* ============ Step 1 control deck ============ */
    .grp{ font-family:'JetBrains Mono',monospace; font-size:.62rem; letter-spacing:.18em;
      text-transform:uppercase; color:var(--muted); font-weight:600;
      margin:18px 0 8px; padding-bottom:5px; border-bottom:1px solid var(--line); }

    .ribbon{ display:flex; flex-wrap:wrap; gap:7px; margin:2px 0 4px; }
    .ribbon .rb{ font-family:'JetBrains Mono',monospace; font-size:.68rem; color:var(--muted);
      background:#F4F7FC; border:1px solid var(--line); border-radius:999px; padding:4px 11px;
      white-space:nowrap; }
    .ribbon .rb b{ color:var(--ink); font-weight:700; }

    /* warehouse node cards */
    .node{ border:1px solid var(--line); border-radius:12px; padding:11px 13px 10px;
      background:var(--surface); position:relative; overflow:hidden; margin-top:-6px;
      transition:opacity .2s ease, border-color .2s ease, transform .2s ease; }
    .node::before{ content:''; position:absolute; left:0; top:0; bottom:0; width:3px;
      background:var(--nc); }
    .node:hover{ transform:translateY(-2px); border-color:var(--nc); }
    .node.off{ opacity:.34; filter:grayscale(1); }
    .node-code{ font-family:'JetBrains Mono',monospace; font-size:.74rem; color:var(--ink);
      font-weight:700; letter-spacing:.04em; }
    .node-serves{ font-family:'JetBrains Mono',monospace; font-size:.62rem; color:var(--nc);
      letter-spacing:.09em; margin-top:4px; font-weight:600; }
    .node-meta{ font-size:.66rem; color:var(--muted); margin-top:3px; }

    /* routing lanes */
    .lanes{ display:flex; flex-direction:column; gap:5px; }
    .lane-row{ display:flex; align-items:center; gap:8px; padding:6px 10px;
      border:1px solid var(--line); border-radius:10px; background:var(--surface);
      transition:border-color .2s ease, background .2s ease; }
    .lane-row:hover{ border-color:#CBD6E8; background:#FBFCFE; }
    .lane-row.broken{ background:#FDF3F3; border-color:#F2C9C9; }
    .ctry{ font-family:'JetBrains Mono',monospace; font-size:.7rem; font-weight:700;
      color:var(--ink); width:38px; flex:none; }
    .ctry-full{ font-size:.75rem; color:var(--muted); width:132px; flex:none;
      overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
    .rail{ flex:1; height:1px; background:repeating-linear-gradient(90deg,
      var(--line) 0 5px, transparent 5px 10px); min-width:14px; }
    .hop{ font-family:'JetBrains Mono',monospace; font-size:.65rem; font-weight:600;
      color:var(--nc); border:1px solid var(--nc); background:color-mix(in srgb, var(--nc) 8%, white);
      border-radius:999px; padding:3px 10px; white-space:nowrap; }
    .hop.bk{ border-style:dashed; opacity:.8; }
    .hop.dead{ color:#B6BFCE; border-color:#DFE4EC; background:#F6F8FB;
      text-decoration:line-through; }
    .arw{ color:#C3CCDC; font-size:.85rem; }
    .warn{ font-family:'JetBrains Mono',monospace; font-size:.6rem; color:var(--rose);
      background:#FBE7E7; border-radius:999px; padding:3px 9px; letter-spacing:.06em; }

    /* rule strips (pack rounding + default target) */
    .rules{ display:flex; flex-direction:column; gap:9px; margin-top:4px; }
    .rule{ border:1px solid var(--line); border-radius:11px; padding:10px 13px;
      background:var(--surface); }
    .rule-head{ display:flex; align-items:baseline; gap:9px; margin-bottom:8px; }
    .rule-pct{ font-family:'Space Grotesk',sans-serif; font-size:1.15rem; font-weight:700;
      color:var(--ink); font-variant-numeric:tabular-nums; }
    .rule-cs{ font-family:'JetBrains Mono',monospace; font-size:.63rem; color:var(--muted);
      letter-spacing:.1em; }
    .strip{ display:flex; align-items:center; gap:4px; flex-wrap:wrap; }
    .strip-cap{ font-family:'JetBrains Mono',monospace; font-size:.61rem; color:var(--muted);
      margin-left:7px; }

    /* proportional threshold bar - one pack, drawn without implying a size */
    .tbar{ display:flex; height:12px; border-radius:6px; overflow:hidden;
      background:#EDF1F7; border:1px solid #E2E8F2; }
    .tb-dn{ background:repeating-linear-gradient(45deg,#DCE3EE 0 4px,#E9EEF6 4px 8px);
      transition:width .28s cubic-bezier(.22,.68,.35,1); }
    .tb-up{ background:linear-gradient(90deg,var(--teal),var(--accent));
      transition:width .28s cubic-bezier(.22,.68,.35,1); }
    .tb-legend{ display:flex; gap:14px; flex-wrap:wrap; margin-top:7px;
      font-family:'JetBrains Mono',monospace; font-size:.6rem; color:var(--muted); }
    .tb-legend .sw{ display:inline-block; width:8px; height:8px; border-radius:2px;
      margin-right:5px; vertical-align:-1px; }
    .tb-legend .sw.dn{ background:#DCE3EE; }
    .tb-legend .sw.up{ background:linear-gradient(90deg,var(--teal),var(--accent)); }
    .tb-ex{ font-family:'JetBrains Mono',monospace; font-size:.58rem; color:var(--muted);
      margin-top:5px; letter-spacing:.02em; }

    /* reserve vault */
    .vault{ border:1px solid var(--line); border-radius:12px; padding:12px 14px 11px;
      background:var(--surface); position:relative; overflow:hidden; }
    .vault::before{ content:''; position:absolute; left:0; top:0; bottom:0; width:3px;
      background:#8B5CF6; }
    .vault.off{ opacity:.45; filter:grayscale(1); }
    .vault-head{ display:flex; align-items:center; justify-content:space-between;
      margin-bottom:10px; }
    .vault-name{ font-family:'JetBrains Mono',monospace; font-size:.72rem; font-weight:700;
      color:var(--ink); letter-spacing:.04em; }
    .vault-tag{ font-family:'JetBrains Mono',monospace; font-size:.6rem; letter-spacing:.08em;
      text-transform:uppercase; padding:3px 9px; border-radius:999px;
      background:#F1F4F9; color:var(--muted); }
    .vault-tag.on{ background:#EFE9FE; color:#6D3FE0; }
    .vault-tag.off-tag{ background:#E6F6F4; color:#0B7C77; }
    .split{ display:flex; height:14px; border-radius:7px; overflow:hidden;
      background:#EDF1F7; }
    .sp-work{ background:linear-gradient(90deg,var(--teal),var(--accent));
      transition:width .3s cubic-bezier(.22,.68,.35,1); }
    .sp-res{ background:repeating-linear-gradient(45deg,#8B5CF6 0 5px,#A78BFA 5px 10px);
      transition:width .3s cubic-bezier(.22,.68,.35,1); }
    .split-legend{ display:flex; gap:16px; flex-wrap:wrap; margin-top:8px;
      font-family:'JetBrains Mono',monospace; font-size:.62rem; color:var(--muted); }
    .split-legend .sw{ display:inline-block; width:9px; height:9px; border-radius:3px;
      margin-right:6px; vertical-align:-1px; }
    .sw.work{ background:linear-gradient(90deg,var(--teal),var(--accent)); }
    .sw.res{ background:#8B5CF6; }
    .vault-note{ font-size:.7rem; color:var(--muted); margin-top:8px; line-height:1.45; }

    /* asymmetric adjustment + cover-day bars */
    .skew{ display:flex; flex-direction:column; gap:6px; margin:10px 0 4px; }
    .skew-row{ display:flex; align-items:center; gap:8px; }
    .skew-lab{ font-family:'JetBrains Mono',monospace; font-size:.6rem; color:var(--muted);
      width:88px; flex:none; text-align:right; }
    .skew-bar{ flex:1; height:10px; border-radius:5px; background:#EDF1F7;
      display:flex; overflow:hidden; }
    .sk-dn{ background:linear-gradient(90deg,#F0A5A5,var(--rose));
      transition:width .28s cubic-bezier(.22,.68,.35,1); }
    .sk-up{ background:linear-gradient(90deg,var(--teal),var(--accent));
      transition:width .28s cubic-bezier(.22,.68,.35,1); }
    .skew-val{ font-family:'JetBrains Mono',monospace; font-size:.66rem; font-weight:700;
      color:var(--ink); width:44px; flex:none; }
    .mirror{ margin:12px 0 4px; }
    .mir-track{ position:relative; height:14px; border-radius:7px; background:#EDF1F7;
      overflow:hidden; }
    .mir-dn{ position:absolute; top:0; bottom:0;
      background:linear-gradient(90deg,var(--rose),#F0A5A5);
      transition:width .28s cubic-bezier(.22,.68,.35,1), left .28s cubic-bezier(.22,.68,.35,1); }
    .mir-up{ position:absolute; top:0; bottom:0;
      background:linear-gradient(90deg,var(--teal),var(--accent));
      transition:width .28s cubic-bezier(.22,.68,.35,1); }
    .mir-axis{ position:absolute; left:50%; top:-2px; bottom:-2px; width:2px;
      background:var(--ink); border-radius:2px; }
    .mir-labs{ display:flex; justify-content:space-between; margin-top:7px;
      font-family:'JetBrains Mono',monospace; font-size:.62rem; }
    .mir-labs i{ font-style:normal; color:var(--muted); font-size:.56rem; }
    .mir-l{ color:var(--rose); font-weight:700; text-align:left; }
    .mir-c{ color:var(--ink); font-weight:700; text-align:center; }
    .mir-r{ color:var(--accent); font-weight:700; text-align:right; }

    .covs{ display:flex; flex-direction:column; gap:4px; margin-top:12px; }
    .cov-row{ display:flex; align-items:center; gap:8px; }
    .cov-c{ font-family:'JetBrains Mono',monospace; font-size:.6rem; font-weight:700;
      color:var(--ink); width:34px; flex:none; }
    .cov-bar{ flex:1; height:7px; border-radius:4px; background:#EDF1F7; overflow:hidden; }
    .cov-bar > span{ display:block; height:100%; border-radius:4px;
      background:linear-gradient(90deg,var(--teal),var(--accent)); }
    .cov-v{ font-family:'JetBrains Mono',monospace; font-size:.6rem; color:var(--muted);
      width:34px; flex:none; }

    /* ---------- sign-in ---------- */
    .auth-wrap{ display:flex; justify-content:center; margin:6vh 0 0; }
    .auth-card{ width:100%; border:1px solid var(--line); border-radius:16px;
      background:var(--surface); padding:22px 24px 18px;
      box-shadow:0 1px 2px rgba(12,18,34,.05), 0 18px 40px rgba(12,18,34,.10); }
    .auth-brand{ display:flex; align-items:center; gap:9px; margin-bottom:16px; }
    .auth-dot{ width:9px; height:9px; border-radius:50%;
      background:linear-gradient(180deg,var(--teal),var(--accent)); }
    .auth-name{ font-family:'JetBrains Mono',monospace; font-size:.68rem;
      letter-spacing:.16em; color:var(--muted); font-weight:600; }
    .auth-steps{ display:flex; align-items:center; gap:10px; }
    .as{ font-family:'JetBrains Mono',monospace; font-size:.62rem; letter-spacing:.1em;
      text-transform:uppercase; color:#B6BFCE; background:#F4F7FC;
      border:1px solid var(--line); border-radius:999px; padding:4px 11px; }
    .as.on{ color:var(--accent); border-color:var(--accent); background:#F2F6FF; }
    .as-rail{ flex:1; height:1px; background:repeating-linear-gradient(90deg,
      var(--line) 0 5px, transparent 5px 10px); }
    .auth-lead{ font-size:.9rem; color:var(--ink-2); margin:16px 0 2px; }

    /* signed-in strip */
    .who{ font-family:'JetBrains Mono',monospace; font-size:.66rem; color:var(--muted);
      display:flex; align-items:center; gap:8px; margin:-14px 0 14px; }
    .who b{ color:var(--ink); }
    .who-dot{ width:7px; height:7px; border-radius:50%; background:var(--teal); }

    /* refill form cards */
    .forms{ display:flex; gap:10px; }
    .form-card{ flex:1; border:1px solid var(--line); border-radius:12px;
      padding:11px 13px 10px; background:var(--surface);
      transition:opacity .2s ease, border-color .2s ease, transform .2s ease; }
    .form-card.on{ border-color:var(--accent); box-shadow:0 6px 18px rgba(47,111,237,.12); }
    .form-card.off{ opacity:.32; filter:grayscale(1); }
    .form-top{ display:flex; align-items:center; justify-content:space-between; }
    .form-name{ font-family:'JetBrains Mono',monospace; font-size:.72rem; font-weight:700;
      color:var(--ink); letter-spacing:.04em; }
    .ord{ font-family:'JetBrains Mono',monospace; font-size:.58rem; font-weight:700;
      color:#fff; background:var(--accent); border-radius:999px; padding:2px 7px; }
    .glyphs{ display:flex; align-items:center; gap:4px; margin:9px 0 7px; min-height:13px; }
    .gl-dot{ width:9px; height:9px; border-radius:50%;
      background:linear-gradient(180deg,var(--teal),var(--accent)); }
    .gl-box{ width:20px; height:13px; border-radius:3px;
      background:linear-gradient(180deg,var(--teal),var(--accent));
      box-shadow:inset 0 0 0 1px rgba(255,255,255,.45); }
    .form-sub{ font-size:.66rem; color:var(--muted); }

    /* Streamlit puts a class named st-key-PLUS-THE-WIDGET-KEY on each widget
       tab buttons can be styled without touching every other button. */
    .st-key-pick_Pack, .st-key-pick_Loose{ margin-top:-30px !important; }
    .st-key-pick_Pack button, .st-key-pick_Loose button{
      border-top-left-radius:0 !important; border-top-right-radius:0 !important;
      border-top:0 !important; min-height:34px; font-size:.78rem;
      font-family:'JetBrains Mono',monospace; letter-spacing:.06em; }
    .st-key-pick_Pack button:disabled, .st-key-pick_Loose button:disabled{
      color:var(--accent) !important; background:#F2F6FF !important;
      border:1px solid var(--accent) !important; border-top:0 !important;
      opacity:1 !important; }

    /* card acts as a tab: flat bottom so it joins its click target */
    .form-card.tabbed{ border-bottom-left-radius:0; border-bottom-right-radius:0;
      border-bottom:none; padding:11px 13px 8px; margin-bottom:0; cursor:default; }
    .form-card.tabbed.on{ border-color:var(--accent); border-bottom:none;
      box-shadow:0 -4px 16px rgba(47,111,237,.10); }
    .form-card.tabbed.off{ opacity:.42; }
    .form-card.tabbed.off:hover{ opacity:.72; border-color:#C7D3E6; }

    /* Pull the tab's click target flush against its card. Streamlit adds a
       st-key-PLUS-THE-WIDGET-KEY class to keyed widgets, so this is scoped
       refill tabs only and leaves every other button alone. If that class is
       ever absent the rule simply does not apply and the layout still works. */
    div[class*="st-key-pick_"]{ margin-top:-1.05rem !important; }
    div[class*="st-key-pick_"] button{
      border-top-left-radius:0 !important; border-top-right-radius:0 !important;
      border-top:none !important; margin-top:0 !important;
      min-height:34px !important; font-size:.78rem !important; }
    </style>
    """)


def render_hero():
    st.html("""
    <div class="hero">
      <div class="hero-eyebrow">Multi-Node &middot; Allocation &middot; Oracle Export</div>
      <div class="hero-title">GCC Replenishment Engine</div>
      <p class="hero-sub">Sales-ranked allocation across the source warehouse network, and an optional reserve held back at JAFZA.</p>
      <div class="lane">
        <span class="lane-node">DIP</span><span class="lane-node">KSA</span>
        <span class="lane-node">QAT</span><span class="lane-node">JAFZA</span>
        <span class="lane-dot"></span><span class="lane-track"></span><span class="lane-dot"></span>
        <span class="lane-node">6 GCC MARKETS</span>
      </div>
    </div>
    """)


def kpi(col, eyebrow, value, label, caption="", tone="", delay="d1", meter=None):
    m = f'<div class="meter"><span style="width:{max(0,min(100,meter)):.1f}%"></span></div>' if meter is not None else ""
    c = f'<div class="kpi-cap">{caption}</div>' if caption else ""
    with col:
        st.html(f"""<div class="kpi {tone} rise {delay}">
          <div class="kpi-eyebrow">{eyebrow}</div><div class="kpi-value">{value}</div>
          <div class="kpi-label">{label}</div>{c}{m}</div>""")


def render_kpis(s, runtime):
    r1 = st.columns(4)
    kpi(r1[0], "Transfer Volume", f"{s['units']:,}", "Units to ship", delay="d1")
    kpi(r1[1], "Order Lines", f"{s['lines']:,}", "Store x item lines", tone="teal", delay="d2")
    kpi(r1[2], "Coverage", f"{s['items']:,}", "Items covered",
        caption=f"across {s['stores']:,} stores", tone="amber", delay="d3")
    kpi(r1[3], "Demand Fill Rate", f"{s['fill']:.1f}%", "Of total rounded need",
        caption=f"{s['got']:,} of {s['need']:,} units", tone="teal", delay="d4", meter=s["fill"])
    r2 = st.columns(4)
    kpi(r2[0], "Reserve Carried Forward", f"{s['reserve']:,}", "Units held at JAFZA",
        caption="secured for the next run", tone="amber", delay="d4")
    kpi(r2[1], "Fallback Sourced", f"{s['fallback']:,}", "Lines from a backup WH",
        caption="priority-1 warehouse could not cover", tone="rose", delay="d5")
    kpi(r2[2], "Unmet Demand", f"{s['unmet_units']:,}", "Units with no stock",
        caption=f"{s['unfilled']:,} store-item lines", tone="rose", delay="d6")
    kpi(r2[3], "Run Time", f"{runtime:.1f}s", "Engine execution",
        caption="excludes file reading", tone="teal", delay="d6")

    if s["defaulted"]:
        st.html(f"""<div class="alert rise d6">
          <b>{s['defaulted']:,} option x store combinations sold but have no row in the DEPTH
          file</b>, so the default target stock grid was applied to size them. They contributed
          <b>{s['defaulted_units']:,} units</b> to this plan. Check the Target Source column in
          the audit trail to see which lines these are.</div>""")


# ---------------------------------------------------------------------------
# APP
# ---------------------------------------------------------------------------
st.set_page_config(page_title="GCC Replenishment Engine", page_icon="\U0001F4E6", layout="wide")
inject_theme()

# Nothing below renders until the visitor is authenticated.
CURRENT_USER = require_login()

render_hero()

_ident, _out = st.columns([5, 1])
with _ident:
    st.html(f'<div class="who"><span class="who-dot"></span>'
            f'Signed in as <b>{CURRENT_USER}</b></div>')
with _out:
    if st.button("Sign out", use_container_width=True):
        for _k in ("auth_user", "auth_stage", "auth_email"):
            st.session_state.pop(_k, None)
        st.rerun()

for k in ("final", "detail", "wh_summary", "summary", "plan_bytes",
          "detail_bytes", "runtime", "read_time"):
    st.session_state.setdefault(k, None)

# --- Step 1: rules --------------------------------------------------------
WH_ACCENT = {"44324": "#2F6FED", "22748": "#0EA5A0", "170001": "#E0A008", "242211": "#8B5CF6"}
COUNTRY_SHORT = {"Kuwait": "KUW", "Qatar": "QAT", "Bahrain": "BAH",
                 "Oman": "OMN", "Saudi Arabia": "KSA", "United Arab Emirates": "UAE"}


# Illustrative pack sizes only - the real size always comes from each
# option's WMS Pack Ratio. These are shown purely so a planner can see that
# the same threshold means a different unit cut-off per item.
EXAMPLE_PACK_SIZES = (9, 10, 12)


def threshold_bar(threshold_pct):
    """
    One pack drawn as a proportional bar, independent of pack size: the left
    portion is the remainder that rounds DOWN, the right portion rounds UP.
    A worked cut-off is shown for a few example pack sizes so it is obvious
    the figure moves with the item's own Pack Ratio.
    """
    down = max(0.0, min(100.0, float(threshold_pct)))
    up = 100.0 - down
    ex = " &middot; ".join(
        f"{n}&rarr;{n * down / 100:.1f}+" for n in EXAMPLE_PACK_SIZES)
    return (f'<div class="tbar">'
            f'<span class="tb-dn" style="width:{down}%"></span>'
            f'<span class="tb-up" style="width:{up}%"></span></div>'
            f'<div class="tb-legend">'
            f'<span><i class="sw dn"></i>below {down:.0f}% of a pack &rarr; rounds down</span>'
            f'<span><i class="sw up"></i>at or above &rarr; rounds up</span></div>'
            f'<div class="tb-ex">cut-off by pack size &nbsp;{ex}</div>')


with st.container(border=True):
    st.html('<div class="step">Step 1 &mdash; Allocation rules</div>')
    st.markdown('<p class="sec-title">Control deck</p>', unsafe_allow_html=True)
    st.markdown('<p class="sec-sub">Everything below has a working default. Touch it only '
                'to override this run.</p>', unsafe_allow_html=True)

    ribbon = st.empty()

    # ---- refill form (pack / loose) ----------------------------------------
    # The two cards ARE the selector - each is a clickable tab. No radio.
    st.html('<div class="grp">Refill form</div>')

    if "refill_mode" not in st.session_state:
        st.session_state.refill_mode = DEFAULT_MODE

    FORM_TABS = [
        (MODE_PACK, "Pack", "whole packs &middot; size from Pack Ratio",
         '<span class="gl-box"></span>' * 3),
        (MODE_LOOSE, "Loose", "eaches &middot; barcode per size",
         '<span class="gl-dot"></span>' * 5),
    ]

    tab_cols = st.columns(len(FORM_TABS))
    for col, (m, label, sub, glyph) in zip(tab_cols, FORM_TABS):
        with col:
            live = st.session_state.refill_mode == m
            st.html(
                f'<div class="form-card tabbed {"on" if live else "off"}">'
                f'<div class="form-top"><span class="form-name">{label}</span>'
                f'{"<span class=ord>active</span>" if live else ""}</div>'
                f'<div class="glyphs">{glyph}</div>'
                f'<div class="form-sub">{sub}</div></div>')
            if st.button("Selected" if live else f"Use {label}",
                         key=f"pick_{m}", use_container_width=True,
                         type="primary" if live else "secondary",
                         disabled=live,
                         help=f"Replenish in {label.lower()}"):
                st.session_state.refill_mode = m
                st.rerun()

    refill_mode = st.session_state.refill_mode
    st.html('<div class="vault-note">' + (
        "Pack size is taken per option from the WMS <b>Pack Ratio</b> column "
        "&mdash; there is no assumed pack of 12. Demand is rounded to that "
        "option's own pack size using the country thresholds below."
        if refill_mode == MODE_PACK else
        "Loose ships in eaches, so no pack rounding is applied &mdash; the "
        "country thresholds below do not affect a loose run.") + '</div>')


    # ---- dynamic target stock by rate of sale --------------------------------
    st.html('<div class="grp">Target stock &mdash; rate of sale</div>')

    ros_on = st.toggle(
        "Adjust target stock by rate of sale", value=ROS_ADJUST_ON,
        help="ROS = lifetime sales / days since first received. Cover Days = SOH / ROS. "
             "A store holding more than its target cover has its target cut; one holding "
             "less has it raised. Needs 'Net Sales Qty' and 'First recieved date store' "
             "in the SOH file.")

    ros_l, ros_r = st.columns([1.25, 1])

    with ros_l:
        cov_df = pd.DataFrame([
            {"Country": c,
             "Cover Days": DEFAULT_COVER_DAYS.get(c, 56),
             "Lead Days": DEFAULT_LEAD_DAYS.get(c, 0)}
            for c in sorted(DEFAULT_COVER_DAYS)])
        if ros_on:
            cov_edit = st.data_editor(
                cov_df, key="cover_editor", hide_index=True, use_container_width=True,
                column_config={
                    "Country": st.column_config.TextColumn("Country", disabled=True),
                    "Cover Days": st.column_config.NumberColumn(
                        "Cover Days", min_value=0, max_value=365, step=7),
                    "Lead Days": st.column_config.NumberColumn(
                        "Lead Days", min_value=0, max_value=90, step=1),
                })
            if st.button("Reset cover days", use_container_width=True):
                st.session_state.pop("cover_editor", None)
                st.rerun()
        else:
            st.dataframe(cov_df, use_container_width=True, hide_index=True)
            cov_edit = cov_df

        cover_days = {}
        for _, rr in cov_edit.iterrows():
            cd = pd.to_numeric(rr["Cover Days"], errors="coerce")
            ld = pd.to_numeric(rr["Lead Days"], errors="coerce")
            cd = 56 if pd.isna(cd) else max(0, float(cd))
            ld = 0 if pd.isna(ld) else max(0, float(ld))
            cover_days[rr["Country"]] = cd + ld

    with ros_r:
        if ros_on:
            adjust_pct = st.slider(
                "Adjustment applied either way", 0, 100, DEFAULT_ADJUST_PCT, 5,
                format="%d%%",
                help="One symmetric figure: an under-covered store's target goes UP "
                     "by this much, an over-covered store's goes DOWN by the same.")
            increase_pct = decrease_pct = adjust_pct

            # Mirrored scale - the centre line is the DEPTH target, the two
            # wings show where the adjusted target lands either side of it.
            w = max(adjust_pct, 1) / 100 * 46
            st.html(
                f'<div class="mirror">'
                f'<div class="mir-track">'
                f'<span class="mir-dn" style="width:{w:.1f}%;left:{50-w:.1f}%"></span>'
                f'<span class="mir-up" style="width:{w:.1f}%;left:50%"></span>'
                f'<span class="mir-axis"></span></div>'
                f'<div class="mir-labs">'
                f'<span class="mir-l">&minus;{adjust_pct}%<br><i>over-covered</i></span>'
                f'<span class="mir-c">DEPTH<br><i>target</i></span>'
                f'<span class="mir-r">+{adjust_pct}%<br><i>under-covered</i></span>'
                f'</div></div>')

            min_ros = st.number_input(
                "Minimum rate of sale (units/day) to replenish at all",
                min_value=0.0, max_value=5.0, value=float(DEFAULT_MIN_ROS),
                step=0.01, format="%.2f",
                help="Lines selling slower than this are excluded. Without a SALE "
                     "file the scope comes from LIFETIME sales, so this is what "
                     "keeps dormant lines out. 0 disables it.")
        else:
            increase_pct = decrease_pct = adjust_pct = 0
            min_ros = 0.0

        rows = "".join(
            f'<div class="cov-row"><span class="cov-c">{COUNTRY_SHORT.get(c, c)}</span>'
            f'<span class="cov-bar"><span style="width:{min(100, v/90*100):.0f}%"></span>'
            f'</span><span class="cov-v">{v:.0f}d</span></div>'
            for c, v in sorted(cover_days.items()))
        st.html(f'<div class="covs">{rows}</div>')
        st.html('<div class="vault-note">' + (
            "Target Cover Days = Cover Days + Lead Days. A store with no rate of sale "
            "keeps its DEPTH target; with neither, the country default applies. Every "
            "line records which rule was used in the audit trail."
            if ros_on else
            "Rate-of-sale adjustment is off &mdash; DEPTH targets are used exactly as "
            "supplied.") + '</div>')


    # ---- warehouse network (editable) --------------------------------------
    if "wh_registry" not in st.session_state:
        st.session_state.wh_registry = dict(GRID_WH)
    if "routing_df" not in st.session_state:
        st.session_state.routing_df = pd.DataFrame(
            routing_rows_from(COUNTRY_PRIORITY, sorted(COUNTRY_PRIORITY)))

    reg = st.session_state.wh_registry
    accent = {c: WH_PALETTE[i % len(WH_PALETTE)] for i, c in enumerate(reg)}

    st.html('<div class="grp">Source warehouses</div>')
    wh_cols = st.columns(max(1, len(reg)))
    active_wh = []
    rdf = st.session_state.routing_df
    for i, (code, name) in enumerate(reg.items()):
        col = wh_cols[i % len(wh_cols)]
        serves, primary = [], []
        if code in rdf.columns:
            for _, r in rdf.iterrows():
                v = pd.to_numeric(r.get(code), errors="coerce")
                if pd.notna(v) and v > 0:
                    serves.append(COUNTRY_SHORT.get(r["Country"], str(r["Country"])[:3].upper()))
                    if v == 1:
                        primary.append(r["Country"])
        with col:
            on = st.checkbox(name, value=True, key=f"wh_{code}")
            st.html(f"""
            <div class="node {'' if on else 'off'}" style="--nc:{accent[code]}">
              <div class="node-code">{code}</div>
              <div class="node-serves">{' '.join(serves) if serves else '&mdash;'}</div>
              <div class="node-meta">{len(primary)} primary &middot; {len(serves)-len(primary)} backup</div>
            </div>""")
        if on:
            active_wh.append(code)

    # ---- routing (editable matrix + live lanes) ----------------------------
    st.html('<div class="grp">Routing</div>')
    st.markdown('<p class="sec-sub">Set a priority number per country: '
                '<b>1</b> is tried first, <b>2</b> next, and a blank cell means that '
                'warehouse cannot serve that market.</p>', unsafe_allow_html=True)

    edit_net = st.toggle("Edit the network", value=False,
                         help="Change priorities, or add and remove warehouses.")

    wh_codes = list(reg.keys())
    for c in wh_codes:
        if c not in st.session_state.routing_df.columns:
            st.session_state.routing_df[c] = None
    keep_cols = ["Country"] + wh_codes
    st.session_state.routing_df = st.session_state.routing_df[
        [c for c in keep_cols if c in st.session_state.routing_df.columns]]

    if edit_net:
        edited = st.data_editor(
            st.session_state.routing_df, hide_index=True, use_container_width=True,
            column_config={
                **{"Country": st.column_config.TextColumn("Country", disabled=True)},
                **{c: st.column_config.NumberColumn(
                    f"{reg[c]}", min_value=1, max_value=9, step=1,
                    help=f"Priority for {reg[c]} ({c}) - blank = not eligible")
                   for c in wh_codes},
            })
        st.session_state.routing_df = edited

        a, b = st.columns(2)
        with a:
            st.markdown("**Add a warehouse**")
            nc = st.text_input("Code", key="new_wh_code", placeholder="e.g. 27405")
            nn = st.text_input("Name", key="new_wh_name", placeholder="e.g. KUWAIT 3PL")
            if st.button("Add warehouse", use_container_width=True):
                nc2, nn2 = str(nc).strip(), str(nn).strip()
                if not nc2 or not nn2:
                    st.warning("Both a code and a name are required.")
                elif nc2 in reg:
                    st.warning(f"{nc2} is already in the network.")
                else:
                    st.session_state.wh_registry[nc2] = nn2
                    st.session_state.routing_df[nc2] = None
                    st.rerun()
        with b:
            st.markdown("**Remove a warehouse**")
            victim = st.selectbox("Warehouse", options=wh_codes,
                                  format_func=lambda c: f"{reg[c]} ({c})",
                                  key="rm_wh")
            st.caption("Any country routed only through it will have no source.")
            if st.button("Remove warehouse", use_container_width=True,
                         disabled=len(reg) <= 1):
                st.session_state.wh_registry.pop(victim, None)
                st.session_state.routing_df = st.session_state.routing_df.drop(
                    columns=[victim], errors="ignore")
                st.session_state.pop(f"wh_{victim}", None)
                st.rerun()

        if st.button("Reset network to default", use_container_width=True):
            st.session_state.wh_registry = dict(GRID_WH)
            st.session_state.routing_df = pd.DataFrame(
                routing_rows_from(COUNTRY_PRIORITY, sorted(COUNTRY_PRIORITY)))
            st.rerun()
    else:
        st.dataframe(st.session_state.routing_df.rename(columns=reg),
                     use_container_width=True, hide_index=True)

    country_priority = priority_from_matrix(st.session_state.routing_df, wh_codes)

    lane_html = ['<div class="lanes">']
    for country in st.session_state.routing_df["Country"]:
        chain = country_priority.get(country, [])
        usable = [w for w in chain if w in active_wh]
        hops = []
        for rank, w in enumerate(chain, start=1):
            live = w in active_wh
            cls = "hop" + ("" if live else " dead") + (" bk" if rank > 1 else "")
            hops.append(f'<span class="{cls}" style="--nc:{accent.get(w,"#7A88A6")}">'
                        f'{reg.get(w, w)}</span>')
            if rank < len(chain):
                hops.append('<span class="arw">&rsaquo;</span>')
        if not chain:
            hops.append('<span class="warn">not routed</span>')
        lane_html.append(
            f'<div class="lane-row{" broken" if not usable else ""}">'
            f'<span class="ctry">{COUNTRY_SHORT.get(country, str(country)[:3].upper())}</span>'
            f'<span class="ctry-full">{country}</span>'
            f'<span class="rail"></span>{"".join(hops)}'
            f'{"<span class=warn>no source</span>" if chain and not usable else ""}</div>')
    lane_html.append('</div>')
    st.html("".join(lane_html))


    # ---- inventory reserve (JAFZA only) ------------------------------------
    st.html('<div class="grp">Inventory reserve</div>')
    rv_ctl, rv_vis = st.columns([1, 1.35])

    with rv_ctl:
        reserve_choice = st.radio(
            f"Reserve inventory from {reg.get(RESERVE_WH_CODE, RESERVE_WH_CODE)}?",
            options=["Yes", "No"],
            index=0 if DEFAULT_RESERVE_ON else 1,
            horizontal=True,
            help="Held back from the working pool and never released during a run. "
                 "It stays in the warehouse and returns in tomorrow's WMS snapshot.",
        )
        reserve_on = reserve_choice == "Yes"
        if reserve_on:
            reserve_pct = st.slider("Reserve %", min_value=0, max_value=50,
                                    value=DEFAULT_RESERVE_PCT, step=1,
                                    format="%d%%", label_visibility="collapsed")
        else:
            reserve_pct = 0

    reserve_map = {RESERVE_WH_CODE: float(reserve_pct)} if reserve_on else {}
    jafza_live = RESERVE_WH_CODE in active_wh and RESERVE_WH_CODE in reg

    with rv_vis:
        work = 100 - reserve_pct
        if not jafza_live:
            vis = ('<div class="vault off"><div class="vault-head">'
                   f'<span class="vault-name">{reg.get(RESERVE_WH_CODE, RESERVE_WH_CODE)}</span>'
                   '<span class="vault-tag">warehouse switched off</span></div>'
                   '<div class="split"><span class="sp-work" style="width:100%"></span></div>'
                   '<div class="split-legend"><span>no stock drawn from this node</span></div>'
                   '</div>')
        elif reserve_on:
            vis = (f'<div class="vault"><div class="vault-head">'
                   f'<span class="vault-name">{reg.get(RESERVE_WH_CODE, RESERVE_WH_CODE)}</span>'
                   f'<span class="vault-tag on">holding {reserve_pct}%</span></div>'
                   f'<div class="split"><span class="sp-work" style="width:{work}%"></span>'
                   f'<span class="sp-res" style="width:{reserve_pct}%"></span></div>'
                   f'<div class="split-legend">'
                   f'<span><i class="sw work"></i>{work}% allocatable now</span>'
                   f'<span><i class="sw res"></i>{reserve_pct}% carried to next run</span>'
                   f'</div></div>')
        else:
            vis = ('<div class="vault"><div class="vault-head">'
                   f'<span class="vault-name">{reg.get(RESERVE_WH_CODE, RESERVE_WH_CODE)}</span>'
                   '<span class="vault-tag off-tag">no reserve</span></div>'
                   '<div class="split"><span class="sp-work" style="width:100%"></span></div>'
                   '<div class="split-legend">'
                   '<span><i class="sw work"></i>100% allocatable now</span></div></div>')
        others = ", ".join(n for c, n in reg.items() if c != RESERVE_WH_CODE) or "No other warehouse"
        st.html(vis + f'<div class="vault-note">{others} allocate their full '
                      f'available quantity &mdash; no reserve applied.</div>')

    # ---- rounding + targets ------------------------------------------------
    c_left, c_right = st.columns([1, 1])

    with c_left:
        st.html('<div class="grp">Pack rounding</div>')
        customise = st.toggle("Override thresholds", value=False,
                              help="A remainder at or above the cut rounds up to a full pack.")
        grid_df = pd.DataFrame([{"Country": c, "Threshold %": p}
                                for c, p in DEFAULT_PACK_THRESHOLD_PCT.items()]
                               ).sort_values("Country").reset_index(drop=True)
        if customise:
            edited = st.data_editor(
                grid_df, key="thr_editor", hide_index=True, use_container_width=True,
                column_config={
                    "Country": st.column_config.TextColumn("Country", disabled=True),
                    "Threshold %": st.column_config.NumberColumn(
                        "Threshold %", min_value=0, max_value=100, step=5, format="%d%%"),
                })
            if st.button("Reset thresholds", use_container_width=True):
                st.session_state.pop("thr_editor", None)
                st.rerun()
        else:
            edited = grid_df

        pack_thresholds = {}
        for _, r in edited.iterrows():
            v = pd.to_numeric(r["Threshold %"], errors="coerce")
            v = FALLBACK_THRESHOLD_PCT if pd.isna(v) else min(100, max(0, float(v)))
            pack_thresholds[r["Country"]] = v

        groups = {}
        for c, p in sorted(pack_thresholds.items()):
            groups.setdefault(p, []).append(COUNTRY_SHORT.get(c, c))
        blocks = []
        for p, cs in sorted(groups.items()):
            blocks.append(f'<div class="rule"><div class="rule-head">'
                          f'<span class="rule-pct">{p:.0f}%</span>'
                          f'<span class="rule-cs">{" ".join(cs)}</span></div>'
                          f'{threshold_bar(p)}</div>')
        st.html(f'<div class="rules">{"".join(blocks)}</div>')

    with c_right:
        st.html('<div class="grp">Default target stock</div>')
        tgt_customise = st.toggle("Override targets", value=False,
                                  help="Used when an Option + Store has no row in DEPTH.")
        tgt_df = pd.DataFrame([{"Country": c, "Default Target Stock": v}
                               for c, v in DEFAULT_TARGET_STOCK.items()]
                              ).sort_values("Country").reset_index(drop=True)
        if tgt_customise:
            tgt_edited = st.data_editor(
                tgt_df, key="tgt_editor", hide_index=True, use_container_width=True,
                column_config={
                    "Country": st.column_config.TextColumn("Country", disabled=True),
                    "Default Target Stock": st.column_config.NumberColumn(
                        "Default Target Stock", min_value=0, max_value=999, step=12),
                })
            if st.button("Reset targets", use_container_width=True):
                st.session_state.pop("tgt_editor", None)
                st.rerun()
        else:
            tgt_edited = tgt_df

        default_targets = {}
        for _, r in tgt_edited.iterrows():
            v = pd.to_numeric(r["Default Target Stock"], errors="coerce")
            v = FALLBACK_TARGET_STOCK if pd.isna(v) else max(0, float(v))
            default_targets[r["Country"]] = v

        tgroups = {}
        for c, v in sorted(default_targets.items()):
            tgroups.setdefault(v, []).append(COUNTRY_SHORT.get(c, c))
        tblocks = []
        vmax = max(tgroups) if tgroups else 1
        for v, cs in sorted(tgroups.items()):
            width = 100.0 * v / vmax if vmax else 0
            tblocks.append(
                f'<div class="rule"><div class="rule-head">'
                f'<span class="rule-pct">{v:.0f}</span>'
                f'<span class="rule-cs">{" ".join(cs)}</span></div>'
                f'<div class="tbar"><span class="tb-up" style="width:{width:.0f}%"></span></div>'
                f'<div class="tb-ex">{v:.0f} units &middot; rounded to each option\'s own '
                f'pack size at run time</div></div>')
        st.html(f'<div class="rules">{"".join(tblocks)}</div>')

    # ---- live config ribbon -------------------------------------------------
    thr_txt = " / ".join(f"{p:.0f}%" for p in sorted(groups))
    tgt_txt = " / ".join(f"{v:.0f}" for v in sorted(tgroups))
    reachable = sum(1 for ch in country_priority.values() if any(w in active_wh for w in ch))
    ribbon.html(f"""
    <div class="ribbon">
      <span class="rb"><b>{len(active_wh)}</b>/{len(reg)} warehouses</span>
      <span class="rb"><b>{reachable}</b>/{len(st.session_state.routing_df)} markets reachable</span>
      <span class="rb">refill <b>{refill_mode}</b></span>
      <span class="rb">thresholds <b>{thr_txt}</b></span>
      <span class="rb">default target <b>{tgt_txt}</b></span>
      <span class="rb">JAFZA reserve <b>{f"{reserve_pct}%" if reserve_on else "off"}</b></span>
      <span class="rb">build <b>{SPEC_VERSION}</b></span>
    </div>""")
# --- Step 2: files --------------------------------------------------------
with st.container(border=True):
    st.html('<div class="step">Step 2 &mdash; Source data</div>')
    st.markdown('<p class="sec-title">Upload the five extracts</p>', unsafe_allow_html=True)
    st.markdown('<p class="sec-sub">DEPTH, SOH and WMS are required. RMS is strongly '
                'recommended (it supplies in-transit stock). SALE is optional &mdash; '
                'supply it to scope the run on a recent period, or omit it and the SOH '
                'lifetime sales are used instead. Files are identified automatically '
                'from their columns, so upload order does not matter.</p>',
                unsafe_allow_html=True)

    uploads = st.file_uploader(
        "Source files (.xlsx or .csv)", type=["xlsx", "xls", "csv", "tsv", "txt"],
        accept_multiple_files=True, label_visibility="collapsed")

    if uploads:
        cache_key = upload_fingerprint(uploads)
        t_read = time.time()
        with st.spinner("Reading and identifying files..."):
            frames, timings, read_warnings = load_sources(uploads, cache_key, SPEC_VERSION)
        read_secs = time.time() - t_read
        st.session_state.read_time = read_secs

        for w in read_warnings:
            st.warning(w)

        cols = st.columns(5)
        for i, kind in enumerate(["DEPTH", "RMS", "SALE", "SOH", "WMS"]):
            with cols[i]:
                if kind in frames:
                    st.metric(kind, f"{len(frames[kind]):,}",
                              help=f"rows parsed in {timings.get(kind,0):.1f}s")
                else:
                    st.metric(kind, "—")
        cap, btn = st.columns([3, 1])
        with cap:
            st.caption(f"Parsed in {read_secs:.1f}s. Files are cached, so re-running does "
                       f"not re-read them. Build {SPEC_VERSION}.")
        with btn:
            if st.button("Re-read files", use_container_width=True,
                         help="Clears the cached parse and reads the uploads again. Use "
                              "this if a column looks missing after an app update."):
                load_sources.clear()
                st.rerun()

        if st.button("Run Replenishment", type="primary", disabled=not frames):
            bar, txt = st.empty(), st.empty()

            def progress_cb(frac, label):
                bar.progress(min(1.0, max(0.0, frac)))
                txt.markdown(
                    f"<span style='font-family:JetBrains Mono,monospace;font-size:.8rem;"
                    f"color:#2F6FED;'>{int(frac*100):3d}%</span> "
                    f"<span style='font-size:.86rem;color:#7A88A6;'>{label}</span>",
                    unsafe_allow_html=True)

            try:
                final, detail, wh_summary, warns, n_defaulted, secs = run_engine(
                    frames, set(active_wh), pack_thresholds, default_targets,
                    reserve_map, refill_mode, country_priority,
                    st.session_state.wh_registry, CURRENT_USER, cover_days,
                    increase_pct, decrease_pct, ros_on, min_ros, progress_cb)

                st.session_state.final = final
                st.session_state.detail = detail
                st.session_state.wh_summary = wh_summary
                st.session_state.summary = summarise(final, detail, wh_summary, n_defaulted)
                st.session_state.runtime = secs
                st.session_state.plan_bytes = to_excel(final, refill_mode)
                st.session_state.detail_bytes = detail.to_csv(index=False).encode("utf-8-sig")

                bar.empty(); txt.empty()
                for w in warns:
                    st.warning(w)
                st.success(f"Replenishment complete - {len(final):,} transfer lines in {secs:.1f}s "
                           f"(files parsed in {read_secs:.1f}s).")
            except ValueError as e:
                for k in ("final", "detail", "wh_summary", "summary"):
                    st.session_state[k] = None
                bar.empty(); txt.empty(); st.error(str(e))
            except Exception as e:
                for k in ("final", "detail", "wh_summary", "summary"):
                    st.session_state[k] = None
                bar.empty(); txt.empty()
                st.error(f"Something went wrong ({type(e).__name__}: {e}).")
                with st.expander("Technical details"):
                    st.code(traceback.format_exc())

# --- Step 3 / 4 -----------------------------------------------------------
if st.session_state.final is not None:
    final = st.session_state.final
    detail = st.session_state.detail
    s = st.session_state.summary

    st.html('<div class="step" style="margin-top:16px;">Step 3 &mdash; Run summary</div>')
    render_kpis(s, st.session_state.runtime)

    with st.container(border=True):
        st.markdown('<p class="sec-title">Warehouse network</p>', unsafe_allow_html=True)
        st.markdown('<p class="sec-sub">What each node shipped, and what it is holding back.</p>',
                    unsafe_allow_html=True)
        st.dataframe(st.session_state.wh_summary, use_container_width=True, hide_index=True)

    with st.container(border=True):
        st.markdown('<p class="sec-title">By country</p>', unsafe_allow_html=True)
        by_c = detail.groupby("Country", as_index=False).agg(
            Lines=("Option", "size"), Need=("Rounded Need", "sum"),
            Allocated=("Allocated Qty", "sum"))
        by_c["Fill %"] = (by_c["Allocated"] / by_c["Need"].replace(0, np.nan) * 100).round(1)
        st.dataframe(by_c, use_container_width=True, hide_index=True)

    st.html('<div class="step" style="margin-top:24px;">Step 4 &mdash; Export</div>')
    with st.container(border=True):
        head, dl = st.columns([3, 1])
        with head:
            st.markdown('<p class="sec-title">Oracle transfer plan</p>', unsafe_allow_html=True)
            st.markdown('<p class="sec-sub">FROM LOCATION and TO LOCATION are location codes '
                        '(Loc Key), ready to import.</p>', unsafe_allow_html=True)
        with dl:
            st.download_button("Download Transfer Plan", data=st.session_state.plan_bytes,
                               file_name="gcc_replenishment_transfer_plan.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               type="primary", use_container_width=True)
        st.dataframe(final.head(5000), use_container_width=True, hide_index=True)
        if len(final) > 5000:
            st.caption(f"Showing the first 5,000 of {len(final):,} lines. The download has all.")

    with st.expander("Calculation detail - audit trail", expanded=False):
        st.markdown('<p class="sec-sub">Every intermediate column behind the allocation: '
                    'country, sales rank, pipeline, failsafes, pack threshold, source '
                    'warehouse and why.</p>', unsafe_allow_html=True)
        st.download_button("Download Audit Trail (CSV)",
                           data=st.session_state.detail_bytes,
                           file_name="gcc_replenishment_audit_trail.csv",
                           mime="text/csv",
                           help="CSV rather than Excel - the audit trail runs to six "
                                "figures of rows and an .xlsx build of that size takes "
                                "about a minute and can exhaust the app's memory. "
                                "Opens directly in Excel.")
        st.dataframe(detail.head(20000), use_container_width=True, hide_index=True)
        if len(detail) > 20000:
            st.caption(f"Showing the first 20,000 of {len(detail):,} rows for page performance.")
